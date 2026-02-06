import requests
import openpyxl
import random
import time
import os
import threading
import asyncio
import re
import tempfile
import json
from pathlib import Path
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters
)
from flask import Flask
from telegram.error import Conflict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import math

# ================== –§–õ–ê–°–ö –î–õ–Ø RENDER ==================
app = Flask(__name__)

@app.route('/')
def home():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Telegram Route Bot</title>
        <meta charset="utf-8">
        <style>
            body {
                font-family: Arial, sans-serif;
                max-width: 800px;
                margin: 0 auto;
                padding: 20px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
            }
            .container {
                background: rgba(255, 255, 255, 0.1);
                backdrop-filter: blur(10px);
                border-radius: 20px;
                padding: 40px;
                box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
                text-align: center;
                border: 1px solid rgba(255, 255, 255, 0.2);
            }
            h1 {
                font-size: 2.5em;
                margin-bottom: 20px;
            }
            .status {
                background: rgba(255, 255, 255, 0.2);
                padding: 15px;
                border-radius: 10px;
                margin: 20px 0;
                font-family: monospace;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>ü§ñ Telegram Route Bot</h1>
            <p>–ë–æ—Ç –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–æ–≤ —É—Å–ø–µ—à–Ω–æ –∑–∞–ø—É—â–µ–Ω!</p>
            <div class="status">
                ‚úÖ –°—Ç–∞—Ç—É—Å: <strong>–ê–ö–¢–ò–í–ï–ù</strong><br>
                üìç –†–µ–∂–∏–º: Web Service<br>
                üöÄ –ü–ª–∞—Ç—Ñ–æ—Ä–º–∞: Render
            </div>
            <p>–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –±–æ—Ç–∞ –≤ Telegram –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–æ–≤</p>
        </div>
    </body>
    </html>
    """

@app.route('/health')
def health():
    return {"status": "ok", "service": "telegram-route-bot"}, 200

def run_flask():
    port = int(os.environ.get('PORT', 10000))
    print(f"üåê Flask —Å–µ—Ä–≤–µ—Ä –∑–∞–ø—É—â–µ–Ω –Ω–∞ –ø–æ—Ä—Ç—É {port}")
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)

# ================== –ù–ê–°–¢–†–û–ô–ö–ò –ë–û–¢–ê ==================
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
GRAPHHOPPER_API_KEY = os.getenv("GRAPHHOPPER_API_KEY", "2c8e643a-360f-47ab-855d-7e884ce217ad")
ORS_API_KEY = os.getenv("ORS_API_KEY", "")  # OpenRouteService API –∫–ª—é—á
YANDEX_GEOCODER_API_KEY = os.getenv("YANDEX_GEOCODER_API_KEY", "")  # –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä API –∫–ª—é—á
USE_ORS_FALLBACK = bool(ORS_API_KEY)
USE_YANDEX_GEOCODER = bool(YANDEX_GEOCODER_API_KEY)

# –§–∏–∫—Å–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –¥–ª—è —Å—Ç–∞—Ä—Ç–æ–≤–æ–π —Ç–æ—á–∫–∏ (–†–æ—Å—Ç–æ–≤-–Ω–∞-–î–æ–Ω—É)
FIXED_START_COORDS = (47.261748, 39.683642)

# ================== –ö–≠–®–ò–†–û–í–ê–ù–ò–ï –ò –õ–û–ì–ò–†–û–í–ê–ù–ò–ï ==================
GEOCODE_CACHE_FILE = "geocode_cache.json"
ROUTE_CACHE_FILE = "route_cache.json"
ERROR_LOG = "errors.log"

def load_geocode_cache():
    """–ó–∞–≥—Ä—É–∂–∞–µ—Ç –∫—ç—à –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è –∏–∑ —Ñ–∞–π–ª–∞"""
    if os.path.exists(GEOCODE_CACHE_FILE):
        try:
            with open(GEOCODE_CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
                print(f"üìÇ –ó–∞–≥—Ä—É–∂–µ–Ω –∫—ç—à –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è: {len(cache)} –∑–∞–ø–∏—Å–µ–π")
                return cache
        except Exception as e:
            print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ –∫—ç—à–∞: {e}")
    return {}

def save_geocode_cache(cache):
    """–°–æ—Ö—Ä–∞–Ω—è–µ—Ç –∫—ç—à –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è –≤ —Ñ–∞–π–ª"""
    try:
        with open(GEOCODE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"üíæ –ö—ç—à —Å–æ—Ö—Ä–∞–Ω–µ–Ω: {len(cache)} –∑–∞–ø–∏—Å–µ–π")
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –∫—ç—à–∞: {e}")

def load_route_cache():
    """–ó–∞–≥—Ä—É–∂–∞–µ—Ç –∫—ç—à –º–∞—Ä—à—Ä—É—Ç–æ–≤ –∏–∑ —Ñ–∞–π–ª–∞"""
    if os.path.exists(ROUTE_CACHE_FILE):
        try:
            with open(ROUTE_CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
                print(f"üìÇ –ó–∞–≥—Ä—É–∂–µ–Ω –∫—ç—à –º–∞—Ä—à—Ä—É—Ç–æ–≤: {len(cache)} –∑–∞–ø–∏—Å–µ–π")
                return cache
        except Exception as e:
            print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ –∫—ç—à–∞ –º–∞—Ä—à—Ä—É—Ç–æ–≤: {e}")
    return {}

def save_route_cache(cache):
    """–°–æ—Ö—Ä–∞–Ω—è–µ—Ç –∫—ç—à –º–∞—Ä—à—Ä—É—Ç–æ–≤ –≤ —Ñ–∞–π–ª"""
    try:
        with open(ROUTE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"üíæ –ö—ç—à –º–∞—Ä—à—Ä—É—Ç–æ–≤ —Å–æ—Ö—Ä–∞–Ω–µ–Ω: {len(cache)} –∑–∞–ø–∏—Å–µ–π")
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –∫—ç—à–∞ –º–∞—Ä—à—Ä—É—Ç–æ–≤: {e}")

def log_error(row_num, address, error_type, details=""):
    """–õ–æ–≥–∏—Ä—É–µ—Ç –æ—à–∏–±–∫–∏ –≤ —Ñ–∞–π–ª"""
    try:
        with open(ERROR_LOG, 'a', encoding='utf-8') as f:
            f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} | Row {row_num} | {error_type} | {address[:100]} | {details}\n")
    except:
        pass

# ================== –§–£–ù–ö–¶–ò–ò –û–ë–†–ê–ë–û–¢–ö–ò –ê–î–†–ï–°–û–í ==================
def clean_text(text):
    """–û—á–∏—Å—Ç–∫–∞ —Ç–µ–∫—Å—Ç–∞ –æ—Ç –ª–∏—à–Ω–∏—Ö —Å–∏–º–≤–æ–ª–æ–≤"""
    if not text:
        return ""
    
    # –ü—Ä–∏–≤–æ–¥–∏–º –∫ —Å—Ç—Ä–æ–∫–µ
    text = str(text)
    
    # –ó–∞–º–µ–Ω—è–µ–º —Ä–∞–∑–ª–∏—á–Ω—ã–µ —Ç–∏–ø—ã —Ç–∏—Ä–µ –Ω–∞ –æ–±—ã—á–Ω—ã–π –¥–µ—Ñ–∏—Å
    text = text.replace('‚Äì', '-').replace('‚Äî', '-').replace('‚àí', '-').replace('‚Äì', '-')
    
    # –ó–∞–º–µ–Ω—è–µ–º —Ç–æ—á–∫–∏ —Å –∑–∞–ø—è—Ç—ã–º–∏ –ø–æ—Å–ª–µ —Å–æ–∫—Ä–∞—â–µ–Ω–∏–π –Ω–∞ –∑–∞–ø—è—Ç—ã–µ
    text = re.sub(r'([–∞-—è–ê-–Ø])\.\s*', r'\1, ', text)
    
    # –£–±–∏—Ä–∞–µ–º –ª–∏—à–Ω–∏–µ –ø—Ä–æ–±–µ–ª—ã
    text = ' '.join(text.split())
    
    # –ó–∞–º–µ–Ω—è–µ–º –¥–≤–æ–π–Ω—ã–µ –¥–µ—Ñ–∏—Å—ã –Ω–∞ –æ–¥–∏–Ω–∞—Ä–Ω—ã–µ
    while '--' in text:
        text = text.replace('--', '-')
    
    # –£–±–∏—Ä–∞–µ–º –ª–∏—à–Ω–∏–µ –∑–∞–ø—è—Ç—ã–µ
    while ',,' in text:
        text = text.replace(',,', ',')
    
    return text.strip()

def normalize_region_name(region):
    """–ù–æ—Ä–º–∞–ª–∏–∑—É–µ—Ç –Ω–∞–∑–≤–∞–Ω–∏–µ —Ä–µ–≥–∏–æ–Ω–∞"""
    if not region:
        return region
    
    region_lower = region.lower()
    
    replacements = {
        "—Ä. –∫–∞—Ä–µ–ª–∏—è": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ö–∞—Ä–µ–ª–∏—è",
        "—Ä. –∫–æ–º–∏": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ö–æ–º–∏",
        "—Ä. –±–∞—à–∫–æ—Ä—Ç–æ—Å—Ç–∞–Ω": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ë–∞—à–∫–æ—Ä—Ç–æ—Å—Ç–∞–Ω",
        "—Ä. –∞–¥—ã–≥–µ—è": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ê–¥—ã–≥–µ—è",
        "—Ä. –º–∞—Ä–∏–π —ç–ª": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ú–∞—Ä–∏–π –≠–ª",
        "—Ä—Å–æ-–∞–ª–∞–Ω–∏—è": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –°–µ–≤–µ—Ä–Ω–∞—è –û—Å–µ—Ç–∏—è-–ê–ª–∞–Ω–∏—è",
        "–∫—á—Ä": "–ö–∞—Ä–∞—á–∞–µ–≤–æ-–ß–µ—Ä–∫–µ—Å—Å–∫–∞—è –†–µ—Å–ø—É–±–ª–∏–∫–∞",
        "–∫–±—Ä": "–ö–∞–±–∞—Ä–¥–∏–Ω–æ-–ë–∞–ª–∫–∞—Ä—Å–∫–∞—è –†–µ—Å–ø—É–±–ª–∏–∫–∞",
        "—Ä. –∫—Ä—ã–º": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ö—Ä—ã–º",
        "—Ä. —Ç–∞—Ç–∞—Ä—Å—Ç–∞–Ω": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –¢–∞—Ç–∞—Ä—Å—Ç–∞–Ω",
        "—Ä. –¥–∞–≥–µ—Å—Ç–∞–Ω": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –î–∞–≥–µ—Å—Ç–∞–Ω",
        "—Ä. –±—É—Ä—è—Ç–∏—è": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ë—É—Ä—è—Ç–∏—è",
        "—Ä. –º–æ—Ä–¥–æ–≤–∏—è": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –ú–æ—Ä–¥–æ–≤–∏—è",
        "—Ä. —É–¥–º—É—Ä—Ç–∏—è": "–£–¥–º—É—Ä—Ç—Å–∫–∞—è –†–µ—Å–ø—É–±–ª–∏–∫–∞",
        "—Ä. —Ö–∞–∫–∞—Å–∏—è": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –•–∞–∫–∞—Å–∏—è",
        "—Ä. —á—É–≤–∞—à–∏—è": "–ß—É–≤–∞—à—Å–∫–∞—è –†–µ—Å–ø—É–±–ª–∏–∫–∞",
        "—Ä. —Å–∞—Ö–∞": "–†–µ—Å–ø—É–±–ª–∏–∫–∞ –°–∞—Ö–∞ (–Ø–∫—É—Ç–∏—è)",
        "–æ–±–ª.": "–æ–±–ª–∞—Å—Ç—å",
        "–∫—Ä–∞–π.": "–∫—Ä–∞–π",
        "—Ä–µ—Å–ø.": "–†–µ—Å–ø—É–±–ª–∏–∫–∞",
        "–∞–≤—Ç.": "–∞–≤—Ç–æ–Ω–æ–º–Ω—ã–π",
        "–∞–æ": "–∞–≤—Ç–æ–Ω–æ–º–Ω—ã–π –æ–∫—Ä—É–≥",
        "—Ä-–Ω": "—Ä–∞–π–æ–Ω",
        "–º–æ": "–º—É–Ω–∏—Ü–∏–ø–∞–ª—å–Ω–æ–µ –æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏–µ",
        "–≥.": "",
        "—Å.": "",
        "–ø.": "",
        "—Å—Ç.": "",
        "—Ö.": "",
        "–¥.": "",
        "—Ä–ø": "—Ä–∞–±–æ—á–∏–π –ø–æ—Å–µ–ª–æ–∫",
        "–ø–≥—Ç": "–ø–æ—Å–µ–ª–æ–∫ –≥–æ—Ä–æ–¥—Å–∫–æ–≥–æ —Ç–∏–ø–∞",
    }
    
    for old, new in replacements.items():
        if old in region_lower:
            region_lower = region_lower.replace(old, new)
    
    # –ö–∞–ø–∏—Ç–∞–ª–∏–∑–∏—Ä—É–µ–º –ø–µ—Ä–≤—É—é –±—É–∫–≤—É –∫–∞–∂–¥–æ–≥–æ —Å–ª–æ–≤–∞
    words = region_lower.split()
    words = [word.capitalize() for word in words if word]
    region = ' '.join(words)
    
    return region

def extract_region_from_address_improved(address):
    """–£–ª—É—á—à–µ–Ω–Ω–æ–µ –∏–∑–≤–ª–µ—á–µ–Ω–∏–µ —Ä–µ–≥–∏–æ–Ω–∞ —Å –∏—Å–ø—Ä–∞–≤–ª–µ–Ω–∏–µ–º –æ–ø–µ—á–∞—Ç–æ–∫"""
    if not address:
        return None
    
    address = clean_text(address)
    
    # –ò—Å–ø—Ä–∞–≤–ª—è–µ–º –æ–ø–µ—á–∞—Ç–∫–∏
    corrections = {
        "—Ä. –∫–≤—Ä–µ–ª–∏—è": "—Ä. –∫–∞—Ä–µ–ª–∏—è",
        "–Ω–∏–∂–µ–≥–æ—Ä–æ–¥–∫—Å–∫–∞—è": "–Ω–∏–∂–µ–≥–æ—Ä–æ–¥—Å–∫–∞—è",
        "–∫–∞–ª—É–±—Å–∫–∞—è": "–∫–∞–ª—É–∂—Å–∫–∞—è",
        "–≤–æ–ª–æ–≥–æ–¥—Å–∫–∞—è –æ–±–ª.": "–≤–æ–ª–æ–≥–æ–¥—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "—Ç–∞–º–±–æ–≤—Å–∫–∞—è –æ–±–ª": "—Ç–∞–º–±–æ–≤—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "–∫–æ—Å—Ç—Ä–æ–º—Å–∫–∞—è –æ–±–ª": "–∫–æ—Å—Ç—Ä–æ–º—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "–∫–∏—Ä–æ–≤—Å–∫–∞—è –æ–±–ª": "–∫–∏—Ä–æ–≤—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "—Å–∞–º–∞—Ä—Å–∫–∞—è –æ–±–ª": "—Å–∞–º–∞—Ä—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "—è—Ä–æ—Å–ª–∞–≤—Å–∫–∞—è –æ–±–ª": "—è—Ä–æ—Å–ª–∞–≤—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "—Ç–≤–µ—Ä—Å–∫–∞—è –æ–±–ª": "—Ç–≤–µ—Ä—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "—Å–º–æ–ª–µ–Ω—Å–∫–∞—è –æ–±–ª": "—Å–º–æ–ª–µ–Ω—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "–Ω–æ–≤–≥–æ—Ä–æ–¥—Å–∫–∞—è –æ–±–ª": "–Ω–æ–≤–≥–æ—Ä–æ–¥—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "–ø—Å–∫–æ–≤—Å–∫–∞—è –æ–±–ª": "–ø—Å–∫–æ–≤—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "–∞—Å—Ç—Ä–∞—Ö–∞–Ω—Å–∫–∞—è –æ–±–ª": "–∞—Å—Ç—Ä–∞—Ö–∞–Ω—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "—Ä–æ—Å—Ç–æ–≤—Å–∫–∞—è –æ–±–ª": "—Ä–æ—Å—Ç–æ–≤—Å–∫–∞—è –æ–±–ª–∞—Å—Ç—å",
        "–∫—Ä–∞—Å–Ω–æ–¥–∞—Ä—Å–∫–∏–π –∫—Ä": "–∫—Ä–∞—Å–Ω–æ–¥–∞—Ä—Å–∫–∏–π –∫—Ä–∞–π",
        "—Å—Ç–∞–≤—Ä–æ–ø–æ–ª—å—Å–∫–∏–π –∫—Ä": "—Å—Ç–∞–≤—Ä–æ–ø–æ–ª—å—Å–∫–∏–π –∫—Ä–∞–π",
        "–∞–ª—Ç–∞–π—Å–∫–∏–π –∫—Ä": "–∞–ª—Ç–∞–π—Å–∫–∏–π –∫—Ä–∞–π",
        "–ø–µ—Ä–º—Å–∫–∏–π –∫—Ä": "–ø–µ—Ä–º—Å–∫–∏–π –∫—Ä–∞–π",
        "–∑–∞–±–∞–π–∫–∞–ª—å—Å–∫–∏–π –∫—Ä": "–∑–∞–±–∞–π–∫–∞–ª—å—Å–∫–∏–π –∫—Ä–∞–π",
        "–∫–∞–º—á–∞—Ç—Å–∫–∏–π –∫—Ä": "–∫–∞–º—á–∞—Ç—Å–∫–∏–π –∫—Ä–∞–π",
    }
    
    address_lower = address.lower()
    for wrong, correct in corrections.items():
        if wrong in address_lower:
            address_lower = address_lower.replace(wrong, correct)
    
    # –ü–∞—Ç—Ç–µ—Ä–Ω—ã –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è —Ä–µ–≥–∏–æ–Ω–∞
    region_patterns = [
        r'^(.*?)\s+(?:–æ–±–ª\.|–æ–±–ª–∞—Å—Ç—å|–∫—Ä–∞–π|—Ä–µ—Å–ø\.|—Ä–µ—Å–ø—É–±–ª–∏–∫–∞|–ê–û|–∞–≤—Ç\.\s+–æ–∫—Ä—É–≥|—Ä-–Ω|—Ä–∞–π–æ–Ω)',
        r'^(—Ä\.\s+[–ê-–Ø–∞-—è—ë–Å\s\-]+)',  # —Ä. –ö–∞—Ä–µ–ª–∏—è
        r'^(?:–ö–ß–†|–ö–ë–†|–†–°–û[\-\s]?–ê–ª–∞–Ω–∏—è|–ß–†|–£–†|–•–ú–ê–û|–Ø–ù–ê–û|–ù–µ–Ω–µ—Ü–∫–∏–π\s+–ê–û)',
        r'^([–ê-–Ø–∞-—è—ë–Å]+\s+[–ê-–Ø–∞-—è—ë–Å]+(?:\s+[–ê-–Ø–∞-—è—ë–Å]+)?)\s+(?:–∫—Ä–∞–π|–æ–±–ª–∞—Å—Ç—å|—Ä–µ—Å–ø—É–±–ª–∏–∫–∞)',
    ]
    
    for pattern in region_patterns:
        match = re.search(pattern, address_lower, re.IGNORECASE)
        if match:
            region = match.group(1).strip()
            if region:
                return normalize_region_name(region)
    
    return None

def extract_settlement_from_address(address):
    """–ò–∑–≤–ª–µ–∫–∞–µ—Ç –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–π –ø—É–Ω–∫—Ç –∏–∑ –∞–¥—Ä–µ—Å–∞"""
    if not address:
        return None
    
    address = clean_text(address)
    
    # –ü–∞—Ç—Ç–µ—Ä–Ω—ã –¥–ª—è –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã—Ö –ø—É–Ω–∫—Ç–æ–≤ —Å —Ä–∞–∑–Ω—ã–º–∏ —Ç–∏–ø–∞–º–∏
    settlement_patterns = [
        # –≥. –ú–æ—Å–∫–≤–∞, –≥.–°–∞–Ω–∫—Ç-–ü–µ—Ç–µ—Ä–±—É—Ä–≥
        r'(?:–≥\.|–≥–æ—Ä–æ–¥\s+|–≥\s+)([^,\-]+)',
        # —Å. –ò–≤–∞–Ω–æ–≤–∫–∞, –ø. –ì–æ—Ä–Ω—ã–π
        r'(?:—Å\.|—Å–µ–ª–æ\s+|–ø\.|–ø–æ—Å—ë–ª–æ–∫\s+|–ø–æ—Å\.|–ø–æ—Å–µ–ª–æ–∫\s+)([^,\-]+)',
        # —Å—Ç-—Ü–∞ –ö–∞–Ω–µ–≤—Å–∫–∞—è, —Å—Ç.–õ–µ–Ω–∏–Ω–≥—Ä–∞–¥—Å–∫–∞—è
        r'(?:—Å—Ç-—Ü–∞\s+|—Å—Ç\.|—Å—Ç–∞–Ω–∏—Ü–∞\s+)([^,\-]+)',
        # –¥. –ü–µ—Ç—Ä–æ–≤–æ, –¥.–ù–æ–≤–æ–µ
        r'(?:–¥\.|–¥–µ—Ä–µ–≤–Ω—è\s+)([^,\-]+)',
        # —Ö. –°–æ–≥–ª–∞—Å–Ω—ã–π
        r'(?:—Ö\.|—Ö—É—Ç–æ—Ä\s+)([^,\-]+)',
        # —Ä.–ø. –ú—É—Ö—Ç–æ–ª–æ–≤–æ
        r'(?:—Ä\.–ø\.|—Ä–∞–±–æ—á–∏–π\s+–ø–æ—Å—ë–ª–æ–∫\s+)([^,\-]+)',
        # –ø–≥—Ç. –ß–µ—Ä–Ω–æ–º–æ—Ä—Å–∫–æ–µ
        r'(?:–ø–≥—Ç\.|–ø–æ—Å—ë–ª–æ–∫\s+–≥–æ—Ä–æ–¥—Å–∫–æ–≥–æ\s+—Ç–∏–ø–∞\s+)([^,\-]+)',
        # –∞—É–ª –ö–æ—à–µ—Ö–∞–±–ª—å
        r'(?:–∞—É–ª\s+)([^,\-]+)',
        # –ï—Å–ª–∏ –µ—Å—Ç—å –∑–∞–ø—è—Ç–∞—è, –±–µ—Ä–µ–º –ø–µ—Ä–≤–æ–µ —Å–ª–æ–≤–æ –¥–æ –∑–∞–ø—è—Ç–æ–π
        r'^[^,]*?,\s*([^,\-]+)(?=,)',
        # –ë–µ—Ä–µ–º –ø–µ—Ä–≤–æ–µ —Å–ª–æ–≤–æ –ø–æ—Å–ª–µ —Ä–µ–≥–∏–æ–Ω–∞
        r'^(?:[–ê-–Ø–∞-—è—ë–Å]+\s+[–ê-–Ø–∞-—è—ë–Å]+(?:\s+[–ê-–Ø–∞-—è—ë–Å]+)?\s+(?:–∫—Ä–∞–π|–æ–±–ª–∞—Å—Ç—å|—Ä–µ—Å–ø—É–±–ª–∏–∫–∞)[,\s]+)?([^,\-]+)',
    ]
    
    for pattern in settlement_patterns:
        match = re.search(pattern, address, re.IGNORECASE)
        if match:
            settlement = match.group(1).strip()
            # –û—á–∏—â–∞–µ–º –æ—Ç –∫–∞–≤—ã—á–µ–∫ –∏ –ª–∏—à–Ω–∏—Ö —Å–∏–º–≤–æ–ª–æ–≤
            settlement = re.sub(r'["¬´¬ª]', '', settlement)
            # –£–±–∏—Ä–∞–µ–º –≤–æ–∑–º–æ–∂–Ω—ã–µ —Ç–æ—á–∫–∏ –≤ –∫–æ–Ω—Ü–µ
            if settlement.endswith('.'):
                settlement = settlement[:-1]
            # –£–±–∏—Ä–∞–µ–º –ª–∏—à–Ω–∏–µ –ø—Ä–æ–±–µ–ª—ã
            settlement = ' '.join(settlement.split())
            return settlement
    
    return None

def parse_address_chain(address_string, default_region=None):
    """–ü–∞—Ä—Å–∏—Ç —Ü–µ–ø–æ—á–∫—É –∞–¥—Ä–µ—Å–æ–≤ —Å —É—á–µ—Ç–æ–º —Ä–µ–≥–∏–æ–Ω–∞ –∏–∑ –ø–µ—Ä–≤–æ–≥–æ –∞–¥—Ä–µ—Å–∞"""
    if not address_string:
        return []
    
    address_string = clean_text(address_string)
    
    # –†–∞–∑–¥–µ–ª—è–µ–º –ø–æ –¥–µ—Ñ–∏—Å—É, –Ω–æ —É—á–∏—Ç—ã–≤–∞–µ–º, —á—Ç–æ –≤ –Ω–∞–∑–≤–∞–Ω–∏—è—Ö –º–æ–≥—É—Ç –±—ã—Ç—å –¥–µ—Ñ–∏—Å—ã
    # –°–Ω–∞—á–∞–ª–∞ –∑–∞–º–µ–Ω—è–µ–º –¥–µ—Ñ–∏—Å—ã –≤ —Å–∫–æ–±–∫–∞—Ö –Ω–∞ –¥—Ä—É–≥–æ–π —Å–∏–º–≤–æ–ª
    temp_char = '¬ß'
    in_brackets = False
    processed = []
    for char in address_string:
        if char == '(':
            in_brackets = True
        elif char == ')':
            in_brackets = False
        if char == '-' and in_brackets:
            processed.append(temp_char)
        else:
            processed.append(char)
    temp_string = ''.join(processed)
    
    # –†–∞–∑–¥–µ–ª—è–µ–º –ø–æ –¥–µ—Ñ–∏—Å–∞–º
    addresses = [addr.replace(temp_char, '-').strip() for addr in re.split(r'\s*-\s*', temp_string) if addr.strip()]
    
    if not addresses:
        return []
    
    # –ò–∑–≤–ª–µ–∫–∞–µ–º —Ä–µ–≥–∏–æ–Ω –∏–∑ –ø–µ—Ä–≤–æ–≥–æ –∞–¥—Ä–µ—Å–∞
    first_region = extract_region_from_address_improved(addresses[0])
    region_to_use = first_region if first_region else default_region
    
    parsed_addresses = []
    
    for i, addr in enumerate(addresses):
        # –ò–∑–≤–ª–µ–∫–∞–µ–º —Ä–µ–≥–∏–æ–Ω –¥–ª—è —Ç–µ–∫—É—â–µ–≥–æ –∞–¥—Ä–µ—Å–∞
        current_region = extract_region_from_address_improved(addr)
        settlement = extract_settlement_from_address(addr)
        
        if not settlement:
            # –ï—Å–ª–∏ –Ω–µ —É–¥–∞–ª–æ—Å—å –∏–∑–≤–ª–µ—á—å –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–π –ø—É–Ω–∫—Ç, –∏—Å–ø–æ–ª—å–∑—É–µ–º –≤–µ—Å—å –∞–¥—Ä–µ—Å
            settlement = addr.split(',')[0] if ',' in addr else addr
        
        # –ï—Å–ª–∏ —É —Ç–µ–∫—É—â–µ–≥–æ –∞–¥—Ä–µ—Å–∞ –Ω–µ—Ç —Ä–µ–≥–∏–æ–Ω–∞, –∏—Å–ø–æ–ª—å–∑—É–µ–º —Ä–µ–≥–∏–æ–Ω –∏–∑ –ø–µ—Ä–≤–æ–≥–æ –∞–¥—Ä–µ—Å–∞
        if not current_region and region_to_use and i > 0:
            # –§–æ—Ä–º–∏—Ä—É–µ–º –ø–æ–ª–Ω—ã–π –∞–¥—Ä–µ—Å —Å —Ä–µ–≥–∏–æ–Ω–æ–º
            full_address = f"{region_to_use}, {settlement}"
        elif current_region:
            full_address = f"{current_region}, {settlement}"
        else:
            full_address = settlement
        
        parsed_addresses.append(full_address)
    
    return parsed_addresses

def extract_all_addresses_from_chain(address_chain):
    """–ò–∑–≤–ª–µ–∫–∞–µ—Ç –≤—Å–µ –∞–¥—Ä–µ—Å–∞ –∏–∑ —Å–ª–æ–∂–Ω–æ–π —Ü–µ–ø–æ—á–∫–∏"""
    if not address_chain:
        return []
    
    # 1. –†–∞–∑–¥–µ–ª—è–µ–º –ø–æ –¥–µ—Ñ–∏—Å–∞–º, –Ω–æ —É—á–∏—Ç—ã–≤–∞–µ–º —Å–ª–æ–∂–Ω—ã–µ —Å–ª—É—á–∞–∏
    addresses = []
    current = ""
    brackets = 0
    
    for char in address_chain:
        if char == '(':
            brackets += 1
        elif char == ')':
            brackets -= 1
        
        if char == '-' and brackets == 0:
            if current.strip():
                addresses.append(current.strip())
            current = ""
        else:
            current += char
    
    if current.strip():
        addresses.append(current.strip())
    
    # 2. –ï—Å–ª–∏ —Ä–∞–∑–¥–µ–ª–µ–Ω–∏–µ –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª–æ, –ø—Ä–æ–±—É–µ–º –¥—Ä—É–≥–∏–µ –º–µ—Ç–æ–¥—ã
    if len(addresses) < 2:
        # –ü—Ä–æ–±—É–µ–º –ø–æ –∑–∞–ø—è—Ç—ã–º
        parts = [p.strip() for p in address_chain.split(',') if len(p.strip()) > 5]
        if len(parts) > 1:
            # –ì—Ä—É–ø–ø–∏—Ä—É–µ–º —á–∞—Å—Ç–∏ –≤ –∞–¥—Ä–µ—Å–∞
            addresses = []
            i = 0
            while i < len(parts):
                if i + 1 < len(parts) and len(parts[i]) < 20:
                    # –û–±—ä–µ–¥–∏–Ω—è–µ–º –∫–æ—Ä–æ—Ç–∫—É—é —á–∞—Å—Ç—å —Å–æ —Å–ª–µ–¥—É—é—â–µ–π
                    addresses.append(f"{parts[i]}, {parts[i+1]}")
                    i += 2
                else:
                    addresses.append(parts[i])
                    i += 1
    
    return addresses

def has_forbidden_region(address):
    """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç, —Å–æ–¥–µ—Ä–∂–∏—Ç –ª–∏ –∞–¥—Ä–µ—Å –∑–∞–ø—Ä–µ—â–µ–Ω–Ω—ã–π —Ä–µ–≥–∏–æ–Ω"""
    if not address:
        return False
    
    forbidden = ['–∫—Ä—ã–º', '–¥–Ω—Ä', '–ª–Ω—Ä', '—Ö–µ—Ä—Å–æ–Ω—Å–∫–∞—è –æ–±–ª', '–∑–∞–ø–æ—Ä–æ–∂—Å–∫–∞—è –æ–±–ª', 
                 '—Å–µ–≤–∞—Å—Ç–æ–ø–æ–ª—å', '–º–∞—Ä–∏—É–ø–æ–ª—å', '–¥–æ–Ω–µ—Ü–∫', '–ª—É–≥–∞–Ω—Å–∫', '–∞–ª—É–ø–∫–∞',
                 '—Ñ–µ–æ–¥–æ—Å–∏—è', '—è–ª—Ç–∞', '—Å–∏–º—Ñ–µ—Ä–æ–ø–æ–ª—å', '–∫–µ—Ä—á—å']
    
    address_lower = address.lower()
    for region in forbidden:
        if region in address_lower:
            return True
    
    return False

def simplify_address_for_geocoding_v2(address):
    """–£–ø—Ä–æ—â–∞–µ—Ç –∞–¥—Ä–µ—Å –¥–ª—è –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è —Å —É—á–µ—Ç–æ–º –æ—Å–æ–±—ã—Ö —Å–ª—É—á–∞–µ–≤"""
    if not address:
        return None
    
    address = clean_text(address)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∑–∞–ø—Ä–µ—â–µ–Ω–Ω—ã–µ —Ä–µ–≥–∏–æ–Ω—ã
    if has_forbidden_region(address):
        print(f"‚ö†Ô∏è –ü—Ä–æ–ø—É—Å–∫–∞—é –∑–∞–ø—Ä–µ—â–µ–Ω–Ω—ã–π —Ä–µ–≥–∏–æ–Ω: {address[:50]}...")
        return None
    
    # –ò—Å–ø—Ä–∞–≤–ª—è–µ–º –æ–ø–µ—á–∞—Ç–∫–∏
    corrections = {
        "—Ä. –∫–≤—Ä–µ–ª–∏—è": "—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –∫–∞—Ä–µ–ª–∏—è",
        "–Ω–∏–∂–µ–≥–æ—Ä–æ–¥–∫—Å–∫–∞—è": "–Ω–∏–∂–µ–≥–æ—Ä–æ–¥—Å–∫–∞—è",
        "–∫–∞–ª—É–±—Å–∫–∞—è": "–∫–∞–ª—É–∂—Å–∫–∞—è",
        "—Ç–≤–µ—Ä—Å–∫–∞—è": "—Ç–≤–µ—Ä—Å–∫–∞—è",
    }
    
    address_lower = address.lower()
    for wrong, correct in corrections.items():
        if wrong in address_lower:
            address_lower = address_lower.replace(wrong, correct)
    
    # –†–∞—Å—à–∏—Ä–µ–Ω–Ω—ã–π —Å–ø–∏—Å–æ–∫ —Ä–µ–≥–∏–æ–Ω–æ–≤ –¥–ª—è –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è
    region_mapping = {
        '—Ä. –∫–∞—Ä–µ–ª–∏—è': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –∫–∞—Ä–µ–ª–∏—è',
        '—Ä. –∫–æ–º–∏': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –∫–æ–º–∏',
        '—Ä. –±–∞—à–∫–æ—Ä—Ç–æ—Å—Ç–∞–Ω': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –±–∞—à–∫–æ—Ä—Ç–æ—Å—Ç–∞–Ω',
        '—Ä. –∞–¥—ã–≥–µ—è': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –∞–¥—ã–≥–µ—è',
        '—Ä. —Ç–∞—Ç–∞—Ä—Å—Ç–∞–Ω': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ —Ç–∞—Ç–∞—Ä—Å—Ç–∞–Ω',
        '—Ä—Å–æ-–∞–ª–∞–Ω–∏—è': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ —Å–µ–≤–µ—Ä–Ω–∞—è –æ—Å–µ—Ç–∏—è-–∞–ª–∞–Ω–∏—è',
        '–∫—á—Ä': '–∫–∞—Ä–∞—á–∞–µ–≤–æ-—á–µ—Ä–∫–µ—Å—Å–∫–∞—è —Ä–µ—Å–ø—É–±–ª–∏–∫–∞',
        '–∫–±—Ä': '–∫–∞–±–∞—Ä–¥–∏–Ω–æ-–±–∞–ª–∫–∞—Ä—Å–∫–∞—è —Ä–µ—Å–ø—É–±–ª–∏–∫–∞',
        '—Ä. –º–æ—Ä–¥–æ–≤–∏—è': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –º–æ—Ä–¥–æ–≤–∏—è',
        '—Ä. –º–∞—Ä–∏–π —ç–ª': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞ –º–∞—Ä–∏–π —ç–ª',
        '—Ä. —É–¥–º—É—Ä—Ç–∏—è': '—É–¥–º—É—Ä—Ç—Å–∫–∞—è —Ä–µ—Å–ø—É–±–ª–∏–∫–∞',
        '—Ä. —á—É–≤–∞—à–∏—è': '—á—É–≤–∞—à—Å–∫–∞—è —Ä–µ—Å–ø—É–±–ª–∏–∫–∞',
        '–æ–±–ª.': '–æ–±–ª–∞—Å—Ç—å',
        '–∫—Ä–∞–π.': '–∫—Ä–∞–π',
        '—Ä–µ—Å–ø.': '—Ä–µ—Å–ø—É–±–ª–∏–∫–∞',
        '–≥.': '',
        '—Å.': '',
        '–ø.': '',
        '—Å—Ç-—Ü–∞': '',
        '—Å—Ç.': '',
        '—Ö.': '',
        '–¥.': '',
        '—Ä–ø.': '',
        '–ø–≥—Ç.': '',
        '–∞—É–ª': '',
    }
    
    # –ó–∞–º–µ–Ω—è–µ–º —Å–æ–∫—Ä–∞—â–µ–Ω–∏—è
    for old, new in region_mapping.items():
        address_lower = address_lower.replace(old, new)
    
    # –í–æ—Å—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –∑–∞–≥–ª–∞–≤–Ω—ã–µ –±—É–∫–≤—ã
    words = address_lower.split()
    words = [w.capitalize() for w in words if w]
    address = ' '.join(words)
    
    # –£–±–∏—Ä–∞–µ–º –ª–∏—à–Ω–∏–µ –∑–∞–ø—è—Ç—ã–µ –∏ –ø—Ä–æ–±–µ–ª—ã
    address = re.sub(r'\s*,\s*', ', ', address)
    address = re.sub(r'\s+', ' ', address)
    
    # –î–æ–±–∞–≤–ª—è–µ–º "Russia" –µ—Å–ª–∏ –Ω–µ—Ç
    if '—Ä–æ—Å—Å–∏—è' not in address.lower() and 'russia' not in address.lower():
        address = f"{address}, –†–æ—Å—Å–∏—è"
    
    return address.strip()

# ================== –ì–ï–û–ö–û–î–ò–†–û–í–ê–ù–ò–ï ==================
def haversine_distance(lat1, lon1, lat2, lon2):
    """–†–∞—Å—Å—Ç–æ—è–Ω–∏–µ –º–µ–∂–¥—É –¥–≤—É–º—è —Ç–æ—á–∫–∞–º–∏ –ø–æ —Ñ–æ—Ä–º—É–ª–µ –≥–∞–≤–µ—Ä—Å–∏–Ω—É—Å–æ–≤ (–≤ –∫–º)"""
    R = 6371  # –†–∞–¥–∏—É—Å –ó–µ–º–ª–∏ –≤ –∫–º
    
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c

def validate_route_distance(distance, coordinates_list):
    """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å —Ä–∞—Å—Å—Ç–æ—è–Ω–∏—è –º–∞—Ä—à—Ä—É—Ç–∞"""
    if not distance or distance <= 0:
        return False
    
    # –ï—Å–ª–∏ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –º–µ–Ω—å—à–µ 10 –∫–º, –ø—Ä–æ–≤–µ—Ä—è–µ–º –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã
    if distance < 10 and len(coordinates_list) >= 2:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ —è–≤–ª—è—é—Ç—Å—è –ª–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –æ–¥–∏–Ω–∞–∫–æ–≤—ã–º–∏
        coord1 = coordinates_list[0]
        for coord2 in coordinates_list[1:]:
            # –í—ã—á–∏—Å–ª—è–µ–º —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –ø–æ –ø—Ä—è–º–æ–π –º–µ–∂–¥—É —Ç–æ—á–∫–∞–º–∏
            lat1, lon1 = coord1
            lat2, lon2 = coord2
            
            # –ü—Ä–∏–º–µ—Ä–Ω–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –ø–æ –ø—Ä—è–º–æ–π (–≤ –∫–º)
            straight_distance = haversine_distance(lat1, lon1, lat2, lon2)
            
            # –ï—Å–ª–∏ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –ø–æ –ø—Ä—è–º–æ–π –±–æ–ª—å—à–µ 50 –∫–º, –∞ –º–∞—Ä—à—Ä—É—Ç –ø–æ–∫–∞–∑—ã–≤–∞–µ—Ç <10 –∫–º, —ç—Ç–æ –æ—à–∏–±–∫–∞
            if straight_distance > 50 and distance < 10:
                print(f"‚ö†Ô∏è –ü–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ: {distance} –∫–º –ø—Ä–∏ –ø—Ä—è–º–æ–π –¥–∏—Å—Ç–∞–Ω—Ü–∏–∏ {straight_distance:.1f} –∫–º")
                return False
    
    return True

def graphhopper_geocode_simple(address, cache):
    """–ü—Ä–æ—Å—Ç–æ–µ –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ —á–µ—Ä–µ–∑ GraphHopper"""
    if not GRAPHHOPPER_API_KEY or not address:
        return None
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫—ç—à
    cache_key = f"gh_{address}"
    if cache_key in cache:
        return cache[cache_key]
    
    url = "https://graphhopper.com/api/1/geocode"
    params = {
        "q": address,
        "key": GRAPHHOPPER_API_KEY,
        "locale": "ru",
        "limit": 1,
        "provider": "default"
    }
    
    try:
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if data.get("hits") and len(data["hits"]) > 0:
                hit = data["hits"][0]
                point = hit.get("point", {})
                lat = point.get("lat")
                lng = point.get("lng")
                
                if lat is not None and lng is not None:
                    coords = (float(lat), float(lng))
                    cache[cache_key] = coords
                    return coords
        
        return None
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ GraphHopper –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è: {e}")
        return None

def yandex_geocode(address, cache):
    """–ì–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ —á–µ—Ä–µ–∑ –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä"""
    if not YANDEX_GEOCODER_API_KEY or not address:
        return None
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫—ç—à
    cache_key = f"yandex_{address}"
    if cache_key in cache:
        return cache[cache_key]
    
    url = "https://geocode-maps.yandex.ru/1.x/"
    params = {
        "apikey": YANDEX_GEOCODER_API_KEY,
        "geocode": address,
        "format": "json",
        "results": 1
    }
    
    try:
        time.sleep(0.1)  # –ü–∞—É–∑–∞ –¥–ª—è —Å–æ–±–ª—é–¥–µ–Ω–∏—è –ª–∏–º–∏—Ç–æ–≤
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            # –ò–∑–≤–ª–µ–∫–∞–µ–º –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã
            try:
                pos = data['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']
                lon, lat = map(float, pos.split())
                coords = (lat, lon)
                cache[cache_key] = coords
                return coords
            except (KeyError, IndexError):
                return None
        else:
            print(f"‚ö†Ô∏è –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä –æ—à–∏–±–∫–∞ {response.status_code}")
            return None
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä–∞: {e}")
        return None

def enhanced_geocode(address, cache):
    """–£–ª—É—á—à–µ–Ω–Ω–æ–µ –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ —Å –Ω–µ—Å–∫–æ–ª—å–∫–∏–º–∏ —Å—Ç—Ä–∞—Ç–µ–≥–∏—è–º–∏"""
    if not address:
        return None
    
    print(f"üìç –ì–µ–æ–∫–æ–¥–∏—Ä—É—é: {address[:60]}...")
    
    # –£–ø—Ä–æ—â–∞–µ–º –∞–¥—Ä–µ—Å
    simplified = simplify_address_for_geocoding_v2(address)
    if not simplified:
        return None
    
    # –°—Ç—Ä–∞—Ç–µ–≥–∏—è 1: GraphHopper
    coords = graphhopper_geocode_simple(simplified, cache)
    if coords:
        print(f"‚úÖ GraphHopper –Ω–∞—à–µ–ª: {coords}")
        return coords
    
    # –°—Ç—Ä–∞—Ç–µ–≥–∏—è 2: –Ø–Ω–¥–µ–∫—Å (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω)
    if USE_YANDEX_GEOCODER:
        coords = yandex_geocode(simplified, cache)
        if coords:
            print(f"‚úÖ –Ø–Ω–¥–µ–∫—Å –Ω–∞—à–µ–ª: {coords}")
            return coords
    
    # –°—Ç—Ä–∞—Ç–µ–≥–∏—è 3: –ü—Ä–æ–±—É–µ–º –±–µ–∑ —Ä–µ–≥–∏–æ–Ω–∞
    settlement = extract_settlement_from_address(address)
    if settlement:
        simple_addr = f"{settlement}, –†–æ—Å—Å–∏—è"
        coords = graphhopper_geocode_simple(simple_addr, cache)
        if coords:
            print(f"‚úÖ GraphHopper –Ω–∞—à–µ–ª (—É–ø—Ä–æ—â–µ–Ω–Ω–æ): {coords}")
            return coords
        
        if USE_YANDEX_GEOCODER:
            coords = yandex_geocode(simple_addr, cache)
            if coords:
                print(f"‚úÖ –Ø–Ω–¥–µ–∫—Å –Ω–∞—à–µ–ª (—É–ø—Ä–æ—â–µ–Ω–Ω–æ): {coords}")
                return coords
    
    print(f"‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞—Ç—å: {address[:50]}...")
    return None

def geocode_start_point(address):
    """–°–ø–µ—Ü–∏–∞–ª—å–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è —Å—Ç–∞—Ä—Ç–æ–≤–æ–π —Ç–æ—á–∫–∏"""
    # –í—Å–µ–≥–¥–∞ –∏—Å–ø–æ–ª—å–∑—É–µ–º —Ñ–∏–∫—Å–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –¥–ª—è –†–æ—Å—Ç–æ–≤–∞-–Ω–∞-–î–æ–Ω—É
    if "—Ä–æ—Å—Ç–æ–≤-–Ω–∞-–¥–æ–Ω—É" in address.lower() or "344064" in address or "–æ–≥–∞–Ω–æ–≤" in address.lower():
        return FIXED_START_COORDS
    
    return enhanced_geocode(address, load_geocode_cache())

# ================== –†–ê–°–ß–ï–¢ –ú–ê–†–®–†–£–¢–û–í ==================
def graphhopper_route_with_waypoints(coordinates_list):
    """–°—Ç—Ä–æ–∏—Ç –º–∞—Ä—à—Ä—É—Ç —á–µ—Ä–µ–∑ –ø—Ä–æ–º–µ–∂—É—Ç–æ—á–Ω—ã–µ —Ç–æ—á–∫–∏ —á–µ—Ä–µ–∑ GraphHopper API"""
    if not GRAPHHOPPER_API_KEY:
        print("‚ö†Ô∏è GRAPHHOPPER_API_KEY –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω!")
        return None
    
    if len(coordinates_list) < 2:
        return None
    
    # ‚ö†Ô∏è GraphHopper –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–µ: –º–∞–∫—Å–∏–º—É–º 4 —Ç–æ—á–∫–∏
    if len(coordinates_list) > 4:
        print(f"‚ö†Ô∏è GraphHopper: —Å–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ —Ç–æ—á–µ–∫ ({len(coordinates_list)}). –ú–∞–∫—Å–∏–º—É–º 4.")
        print("‚ö†Ô∏è –ë—É–¥—É –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ 4 —Ç–æ—á–∫–∏")
        coordinates_list = coordinates_list[:4]
    
    # –°–æ–∑–¥–∞–µ–º –∫–ª—é—á –¥–ª—è –∫—ç—à–∞
    coords_str = '|'.join([f"{lat:.6f},{lon:.6f}" for lat, lon in coordinates_list])
    cache_key = f"gh_route_{coords_str}"
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫—ç—à –º–∞—Ä—à—Ä—É—Ç–æ–≤
    route_cache = load_route_cache()
    if cache_key in route_cache:
        distance = route_cache[cache_key]
        print(f"‚úÖ –ú–∞—Ä—à—Ä—É—Ç –∏–∑ –∫—ç—à–∞: {distance} –∫–º")
        return distance
    
    url = "https://graphhopper.com/api/1/route"
    
    # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –¥–ª—è –∑–∞–ø—Ä–æ—Å–∞
    params = {
        "key": GRAPHHOPPER_API_KEY,
        "vehicle": "car",
        "locale": "ru",
        "instructions": "false",
        "calc_points": "false",
        "points_encoded": "false",
        "elevation": "false",
        "optimize": "false"
    }
    
    # –î–æ–±–∞–≤–ª—è–µ–º —Ç–æ—á–∫–∏ –º–∞—Ä—à—Ä—É—Ç–∞ –≤ —Ñ–æ—Ä–º–∞—Ç–µ "lat,lng"
    points = []
    for i, coord in enumerate(coordinates_list):
        points.append(f"point={coord[0]},{coord[1]}")
    
    # –§–æ—Ä–º–∏—Ä—É–µ–º URL —Å –ø–∞—Ä–∞–º–µ—Ç—Ä–∞–º–∏
    query_string = "&".join(points) + "&" + "&".join([f"{k}={v}" for k, v in params.items()])
    
    try:
        print(f"üìç GraphHopper —Å—Ç—Ä–æ–∏—Ç –º–∞—Ä—à—Ä—É—Ç —á–µ—Ä–µ–∑ {len(coordinates_list)} —Ç–æ—á–µ–∫...")
        
        full_url = f"{url}?{query_string}"
        r = requests.get(full_url, timeout=60)
        
        if r.status_code != 200:
            print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –º–∞—Ä—à—Ä—É—Ç–∞ {r.status_code}")
            
            # –ï—Å–ª–∏ —Ç–æ—á–µ–∫ –±—ã–ª–æ 4 –∏ –æ—à–∏–±–∫–∞ 400, –ø—Ä–æ–±—É–µ–º —Å 3 —Ç–æ—á–∫–∞–º–∏
            if r.status_code == 400 and len(coordinates_list) == 4:
                print("üîÑ –ü—Ä–æ–±—É—é —Å 3 —Ç–æ—á–∫–∞–º–∏...")
                # –ü—Ä–æ–±—É–µ–º –±–µ–∑ –ø—Ä–µ–¥–ø–æ—Å–ª–µ–¥–Ω–µ–π —Ç–æ—á–∫–∏
                new_coords = [coordinates_list[0], coordinates_list[1], coordinates_list[3]]
                return graphhopper_route_with_waypoints(new_coords)
            
            # –ü—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –¥–µ—Ç–∞–ª–∏ –æ—à–∏–±–∫–∏
            try:
                error_details = r.json()
                print(f"‚ö†Ô∏è –î–µ—Ç–∞–ª–∏ –æ—à–∏–±–∫–∏: {error_details}")
                if "Too many points" in str(error_details):
                    print("üîÑ –°–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ —Ç–æ—á–µ–∫, –ø—Ä–æ–±—É—é —É–º–µ–Ω—å—à–∏—Ç—å...")
                    if len(coordinates_list) > 2:
                        return graphhopper_route_with_waypoints(coordinates_list[:len(coordinates_list)-1])
            except:
                print(f"‚ö†Ô∏è –¢–µ–∫—Å—Ç –æ—à–∏–±–∫–∏: {r.text[:200]}")
            return None
        
        data = r.json()
        
        if data.get("paths") and len(data["paths"]) > 0:
            path = data["paths"][0]
            distance_meters = path.get("distance", 0)
            
            if distance_meters > 0:
                distance_km = round(distance_meters / 1000, 1)
                print(f"‚úÖ –ú–∞—Ä—à—Ä—É—Ç –ø–æ—Å—Ç—Ä–æ–µ–Ω: {distance_km} –∫–º")
                
                # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –∫—ç—à
                route_cache[cache_key] = distance_km
                save_route_cache(route_cache)
                
                return distance_km
            else:
                print(f"‚ö†Ô∏è –ù—É–ª–µ–≤–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –≤ –º–∞—Ä—à—Ä—É—Ç–µ")
                return None
        else:
            print(f"‚ö†Ô∏è –ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π –æ—Ç–≤–µ—Ç –æ—Ç GraphHopper")
            return None
            
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ—Å—Ç—Ä–æ–µ–Ω–∏–∏ –º–∞—Ä—à—Ä—É—Ç–∞: {e}")
        return None

def ors_route_with_waypoints(coordinates_list):
    """–°—Ç—Ä–æ–∏—Ç –º–∞—Ä—à—Ä—É—Ç —á–µ—Ä–µ–∑ OpenRouteService API (–∑–∞–ø–∞—Å–Ω–æ–π –≤–∞—Ä–∏–∞–Ω—Ç)"""
    if not ORS_API_KEY:
        print("‚ö†Ô∏è ORS_API_KEY –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω!")
        return None
    
    if len(coordinates_list) < 2:
        return None
    
    # ORS –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç –¥–æ 50 —Ç–æ—á–µ–∫, –Ω–æ –æ–≥—Ä–∞–Ω–∏—á–∏–º 20 –¥–ª—è –Ω–∞–¥–µ–∂–Ω–æ—Å—Ç–∏
    if len(coordinates_list) > 20:
        print(f"‚ö†Ô∏è ORS: —Å–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ —Ç–æ—á–µ–∫ ({len(coordinates_list)}). –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞—é 20.")
        coordinates_list = coordinates_list[:20]
    
    # –°–æ–∑–¥–∞–µ–º –∫–ª—é—á –¥–ª—è –∫—ç—à–∞
    coords_str = '|'.join([f"{lat:.6f},{lon:.6f}" for lat, lon in coordinates_list])
    cache_key = f"ors_route_{coords_str}"
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫—ç—à –º–∞—Ä—à—Ä—É—Ç–æ–≤
    route_cache = load_route_cache()
    if cache_key in route_cache:
        distance = route_cache[cache_key]
        print(f"‚úÖ ORS –º–∞—Ä—à—Ä—É—Ç –∏–∑ –∫—ç—à–∞: {distance} –∫–º")
        return distance
    
    url = "https://api.openrouteservice.org/v2/directions/driving-car"
    
    # ORS –∏—Å–ø–æ–ª—å–∑—É–µ—Ç —Ñ–æ—Ä–º–∞—Ç [–¥–æ–ª–≥–æ—Ç–∞, —à–∏—Ä–æ—Ç–∞]
    coordinates_ors = [[lon, lat] for lat, lon in coordinates_list]
    
    headers = {
        'Authorization': ORS_API_KEY,
        'Content-Type': 'application/json'
    }
    
    body = {
        "coordinates": coordinates_ors,
        "instructions": False,
        "geometry": False,
        "units": "km"
    }
    
    try:
        print(f"üìç ORS —Å—Ç—Ä–æ–∏—Ç –º–∞—Ä—à—Ä—É—Ç —á–µ—Ä–µ–∑ {len(coordinates_list)} —Ç–æ—á–µ–∫...")
        
        r = requests.post(url, json=body, headers=headers, timeout=60)
        
        if r.status_code != 200:
            print(f"‚ö†Ô∏è ORS –æ—à–∏–±–∫–∞ –º–∞—Ä—à—Ä—É—Ç–∞ {r.status_code}")
            print(f"‚ö†Ô∏è –û—Ç–≤–µ—Ç: {r.text[:200]}")
            return None
        
        data = r.json()
        
        if data.get("routes") and len(data["routes"]) > 0:
            route = data["routes"][0]
            distance_km = round(route.get("summary", {}).get("distance", 0) / 1000, 1)
            
            if distance_km > 0:
                print(f"‚úÖ ORS –º–∞—Ä—à—Ä—É—Ç –ø–æ—Å—Ç—Ä–æ–µ–Ω: {distance_km} –∫–º")
                
                # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –∫—ç—à
                route_cache[cache_key] = distance_km
                save_route_cache(route_cache)
                
                return distance_km
            else:
                print(f"‚ö†Ô∏è ORS –Ω—É–ª–µ–≤–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –≤ –º–∞—Ä—à—Ä—É—Ç–µ")
                return None
        else:
            print(f"‚ö†Ô∏è –ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π –æ—Ç–≤–µ—Ç –æ—Ç ORS")
            return None
            
    except Exception as e:
        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ—Å—Ç—Ä–æ–µ–Ω–∏–∏ –º–∞—Ä—à—Ä—É—Ç–∞ –≤ ORS: {e}")
        return None

def calculate_route_segments(coordinates_list):
    """–†–∞–∑–±–∏–≤–∞–µ—Ç –º–∞—Ä—à—Ä—É—Ç —Å –º–Ω–æ–≥–∏–º–∏ —Ç–æ—á–∫–∞–º–∏ –Ω–∞ —Å–µ–≥–º–µ–Ω—Ç—ã –ø–æ 4 —Ç–æ—á–∫–∏"""
    if len(coordinates_list) <= 4:
        # –ü—Ä–æ–±—É–µ–º GraphHopper
        distance = graphhopper_route_with_waypoints(coordinates_list)
        if distance:
            return distance
        
        # –ü—Ä–æ–±—É–µ–º ORS –∫–∞–∫ –∑–∞–ø–∞—Å–Ω–æ–π –≤–∞—Ä–∏–∞–Ω—Ç
        if USE_ORS_FALLBACK:
            distance = ors_route_with_waypoints(coordinates_list)
            if distance:
                return distance
        
        return None
    
    # –î–ª—è –º–∞—Ä—à—Ä—É—Ç–æ–≤ —Å 5-20 —Ç–æ—á–∫–∞–º–∏ —Å–Ω–∞—á–∞–ª–∞ –ø—Ä–æ–±—É–µ–º ORS —Ü–µ–ª–∏–∫–æ–º
    if 5 <= len(coordinates_list) <= 20 and USE_ORS_FALLBACK:
        distance = ors_route_with_waypoints(coordinates_list)
        if distance:
            return distance
    
    # –ï—Å–ª–∏ ORS –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª –∏–ª–∏ —Ç–æ—á–µ–∫ >20, —Ä–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ —Å–µ–≥–º–µ–Ω—Ç—ã
    print(f"üìç –†–∞–∑–±–∏–≤–∞—é –º–∞—Ä—à—Ä—É—Ç –Ω–∞ —Å–µ–≥–º–µ–Ω—Ç—ã ({len(coordinates_list)} —Ç–æ—á–µ–∫)...")
    
    total_distance = 0
    segments = []
    
    # –†–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ —Å–µ–≥–º–µ–Ω—Ç—ã –ø–æ 4 —Ç–æ—á–∫–∏ (—Å—Ç–∞—Ä—Ç + 3 –ø—Ä–æ–º–µ–∂—É—Ç–æ—á–Ω—ã–µ)
    for i in range(0, len(coordinates_list)-1, 3):
        segment = coordinates_list[i:i+4]
        if len(segment) < 2:
            continue
        
        # –ü–æ—Å–ª–µ–¥–Ω—è—è —Ç–æ—á–∫–∞ —Å–µ–≥–º–µ–Ω—Ç–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–µ—Ä–≤–æ–π —Å–ª–µ–¥—É—é—â–µ–≥–æ —Å–µ–≥–º–µ–Ω—Ç–∞
        if i > 0 and segments:
            # –£–±–µ–¥–∏–º—Å—è, —á—Ç–æ –µ—Å—Ç—å –ø–µ—Ä–µ–∫—Ä—ã—Ç–∏–µ
            if segment[0] != segments[-1][-1]:
                segment.insert(0, segments[-1][-1])
        
        segments.append(segment)
    
    # –ï—Å–ª–∏ —Å–µ–≥–º–µ–Ω—Ç–æ–≤ —Å–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ, —É–ø—Ä–æ—â–∞–µ–º
    if len(segments) > 10:
        print(f"‚ö†Ô∏è –°–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ —Å–µ–≥–º–µ–Ω—Ç–æ–≤ ({len(segments)}), —É–ø—Ä–æ—â–∞—é...")
        # –ë–µ—Ä–µ–º —Ç–æ–ª—å–∫–æ –∫–ª—é—á–µ–≤—ã–µ —Ç–æ—á–∫–∏: —Å—Ç–∞—Ä—Ç, 1/4, 1/2, 3/4, –∫–æ–Ω–µ—Ü
        key_indices = [0]
        if len(coordinates_list) > 4:
            key_indices.append(len(coordinates_list) // 4)
        key_indices.append(len(coordinates_list) // 2)
        key_indices.append(3 * len(coordinates_list) // 4)
        key_indices.append(len(coordinates_list) - 1)
        
        key_points = [coordinates_list[i] for i in key_indices]
        return calculate_route_segments(key_points)
    
    # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –∫–∞–∂–¥—ã–π —Å–µ–≥–º–µ–Ω—Ç
    for idx, segment in enumerate(segments):
        print(f"üìç –°–µ–≥–º–µ–Ω—Ç {idx+1}/{len(segments)}: {len(segment)} —Ç–æ—á–µ–∫")
        
        # –ü—Ä–æ–±—É–µ–º GraphHopper –¥–ª—è —Å–µ–≥–º–µ–Ω—Ç–∞
        segment_distance = graphhopper_route_with_waypoints(segment)
        
        # –ï—Å–ª–∏ –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª–æ, –ø—Ä–æ–±—É–µ–º ORS
        if not segment_distance and USE_ORS_FALLBACK:
            segment_distance = ors_route_with_waypoints(segment)
        
        if segment_distance:
            total_distance += segment_distance
        else:
            print(f"‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å—Å—á–∏—Ç–∞—Ç—å —Å–µ–≥–º–µ–Ω—Ç {idx+1}")
            return None
    
    return total_distance

def validate_coordinates(coords_list):
    """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç, —á—Ç–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —Ä–∞–∑—É–º–Ω—ã–µ –¥–ª—è –†–æ—Å—Å–∏–∏"""
    if not coords_list:
        return False
    
    for lat, lon in coords_list:
        # –†–æ—Å—Å–∏—è –ø—Ä–∏–º–µ—Ä–Ω–æ –≤ –ø—Ä–µ–¥–µ–ª–∞—Ö:
        # –®–∏—Ä–æ—Ç–∞: 41¬∞ –¥–æ 82¬∞ N
        # –î–æ–ª–≥–æ—Ç–∞: 19¬∞ –¥–æ 190¬∞ E
        if not (40 <= lat <= 83) or not (19 <= lon <= 191):
            print(f"‚ö†Ô∏è –ü–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω—ã–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã: {lat}, {lon}")
            return False
    
    return True

def calculate_route(coordinates_list):
    """–û—Å–Ω–æ–≤–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–∞ —Å –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ–º –≤—Å–µ—Ö –¥–æ—Å—Ç—É–ø–Ω—ã—Ö –º–µ—Ç–æ–¥–æ–≤"""
    if len(coordinates_list) < 2:
        return None
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤–∞–ª–∏–¥–Ω–æ—Å—Ç—å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç
    if not validate_coordinates(coordinates_list):
        print("‚ö†Ô∏è –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –≤—ã–≥–ª—è–¥—è—Ç –ø–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω–æ")
        return None
    
    # –í–∞–ª–∏–¥–∞—Ü–∏—è: –ø—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –≤—Å–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —Ä–∞–∑–Ω—ã–µ
    unique_coords = set([f"{lat:.4f},{lon:.4f}" for lat, lon in coordinates_list])
    if len(unique_coords) != len(coordinates_list):
        print("‚ö†Ô∏è –û–±–Ω–∞—Ä—É–∂–µ–Ω—ã –¥—É–±–ª–∏—Ä—É—é—â–∏–µ—Å—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã")
        # –£–¥–∞–ª—è–µ–º –¥—É–±–ª–∏–∫–∞—Ç—ã
        seen = set()
        unique_list = []
        for coord in coordinates_list:
            key = f"{coord[0]:.4f},{coord[1]:.4f}"
            if key not in seen:
                seen.add(key)
                unique_list.append(coord)
        
        if len(unique_list) < 2:
            return None
        
        coordinates_list = unique_list
        print(f"üìç –£–¥–∞–ª–µ–Ω—ã –¥—É–±–ª–∏–∫–∞—Ç—ã, –æ—Å—Ç–∞–ª–æ—Å—å {len(coordinates_list)} —Ç–æ—á–µ–∫")
    
    # –ü—Ä–æ–±—É–µ–º —Ä–∞–∑–Ω—ã–µ —Å—Ç—Ä–∞—Ç–µ–≥–∏–∏ —Ä–∞—Å—á–µ—Ç–∞
    strategies = [
        ("GraphHopper —Ü–µ–ª–∏–∫–æ–º", lambda: graphhopper_route_with_waypoints(coordinates_list)),
    ]
    
    if USE_ORS_FALLBACK:
        strategies.append(("ORS —Ü–µ–ª–∏–∫–æ–º", lambda: ors_route_with_waypoints(coordinates_list)))
    
    strategies.append(("–°–µ–≥–º–µ–Ω—Ç–∞—Ä–Ω—ã–π —Ä–∞—Å—á–µ—Ç", lambda: calculate_route_segments(coordinates_list)))
    
    for strategy_name, strategy_func in strategies:
        print(f"üìç –ü—Ä–æ–±—É—é —Å—Ç—Ä–∞—Ç–µ–≥–∏—é: {strategy_name}")
        distance = strategy_func()
        if distance and distance > 0:
            print(f"‚úÖ –£—Å–ø–µ—à–Ω–æ —Å —Å—Ç—Ä–∞—Ç–µ–≥–∏–µ–π: {strategy_name}")
            return distance
    
    print("‚ùå –í—Å–µ —Å—Ç—Ä–∞—Ç–µ–≥–∏–∏ —Ä–∞—Å—á–µ—Ç–∞ –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª–∏")
    return None

def smart_variations(base_distance):
    """–£–º–Ω—ã–µ –≤–∞—Ä–∏–∞—Ü–∏–∏ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–π —Å –ø—Ä–æ–≤–µ—Ä–∫–æ–π –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç–∏"""
    if not base_distance or base_distance <= 0:
        return [None, None]
    
    # –ï—Å–ª–∏ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –º–∞–ª–µ–Ω—å–∫–æ–µ (–º–µ–Ω–µ–µ 10 –∫–º), –Ω–µ –¥–æ–±–∞–≤–ª—è–µ–º –≤–∞—Ä–∏–∞—Ü–∏–∏
    if base_distance < 10:
        return [None, None]
    
    # –ï—Å–ª–∏ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ –ø–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω–æ –º–∞–ª–µ–Ω—å–∫–æ–µ –¥–ª—è –º–µ–∂–¥—É–≥–æ—Ä–æ–¥–Ω–µ–≥–æ –º–∞—Ä—à—Ä—É—Ç–∞
    if base_distance < 50:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ —è–≤–ª—è–µ—Ç—Å—è –ª–∏ —ç—Ç–æ –æ—à–∏–±–∫–æ–π
        return [round(base_distance * 1.02, 1), round(base_distance * 0.98, 1)]
    
    # –ù–æ—Ä–º–∞–ª—å–Ω—ã–µ –≤–∞—Ä–∏–∞—Ü–∏–∏ 2-5%
    variation_percent = random.uniform(0.02, 0.05)
    variation = base_distance * variation_percent
    
    var1 = round(base_distance + random.uniform(variation/2, variation), 1)
    var2 = round(max(base_distance * 0.95, base_distance - random.uniform(variation/2, variation)), 1)
    
    return [var1, var2]

# ================== –ß–¢–ï–ù–ò–ï –ò –ó–ê–ü–ò–°–¨ EXCEL ==================
def read_excel_with_fallback(file_path):
    """–ß–∏—Ç–∞–µ—Ç Excel —Ñ–∞–π–ª —Å –ø–æ–º–æ—â—å—é openpyxl"""
    try:
        print(f"üìñ –ß—Ç–µ–Ω–∏–µ —Ñ–∞–π–ª–∞ —Å openpyxl...")
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # –°–æ–±–∏—Ä–∞–µ–º –¥–∞–Ω–Ω—ã–µ
        data = []
        max_row = ws.max_row
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º, –µ—Å—Ç—å –ª–∏ –∑–∞–≥–æ–ª–æ–≤–∫–∏ (–ø—Ä–æ–≤–µ—Ä—è–µ–º –ø–µ—Ä–≤—É—é —Å—Ç—Ä–æ–∫—É)
        has_headers = False
        if max_row > 0:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–µ—Ä–≤—ã–µ 2 —è—á–µ–π–∫–∏ –ø–µ—Ä–≤–æ–π —Å—Ç—Ä–æ–∫–∏
            cell1 = ws.cell(row=1, column=1).value
            cell2 = ws.cell(row=1, column=2).value
            
            # –ï—Å–ª–∏ –≤ –ø–µ—Ä–≤–æ–π —Å—Ç—Ä–æ–∫–µ –µ—Å—Ç—å —Å–ª–æ–≤–∞ "–ø—É–Ω–∫—Ç", "–Ω–∞–∑–Ω–∞—á–µ–Ω–∏–µ" –∏ —Ç.–¥., —Ç–æ —ç—Ç–æ –∑–∞–≥–æ–ª–æ–≤–∫–∏
            if cell1 and cell2:
                text1 = str(cell1).lower()
                text2 = str(cell2).lower()
                header_keywords = ['–ø—É–Ω–∫—Ç', '–Ω–∞–∑–Ω–∞—á–µ–Ω–∏–µ', '–≥—Ä—É–∑', '–∞–¥—Ä–µ—Å', '—Ç–æ—á–∫–∞', '–æ—Ç–ø—Ä–∞–≤', '–ø–æ–ª—É—á']
                has_headers = any(keyword in text1 for keyword in header_keywords) or \
                             any(keyword in text2 for keyword in header_keywords)
        
        start_row = 2 if has_headers else 1
        
        for row in range(start_row, max_row + 1):
            col1 = ws.cell(row=row, column=1).value
            col2 = ws.cell(row=row, column=2).value
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –æ–±–µ —è—á–µ–π–∫–∏ –Ω–µ –ø—É—Å—Ç—ã–µ
            if col1 is not None and col2 is not None:
                start_point = clean_text(str(col1))
                address_chain = clean_text(str(col2))
                
                # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º —Å—Ç—Ä–æ–∫–∏, –≥–¥–µ —Å–ª–∏—à–∫–æ–º –º–∞–ª–æ —Å–∏–º–≤–æ–ª–æ–≤
                if len(start_point) > 3 and len(address_chain) > 3:
                    data.append({
                        'row_num': row,
                        'start_point': start_point,
                        'address_chain': address_chain,
                        'original_start': col1,
                        'original_chain': col2
                    })
        
        print(f"‚úÖ –£—Å–ø–µ—à–Ω–æ –ø—Ä–æ—á–∏—Ç–∞–Ω–æ {len(data)} —Å—Ç—Ä–æ–∫")
        return data, wb, ws
        
    except Exception as e:
        print(f"‚ùå –û—à–∏–±–∫–∞ —á—Ç–µ–Ω–∏—è —Ñ–∞–π–ª–∞: {e}")
        raise Exception(f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ—á–∏—Ç–∞—Ç—å —Ñ–∞–π–ª. –£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ —ç—Ç–æ –∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π Excel —Ñ–∞–π–ª (—Ñ–æ—Ä–º–∞—Ç .xlsx). –û—à–∏–±–∫–∞: {str(e)[:200]}")

def add_result_columns(ws, start_col=3):
    """–î–æ–±–∞–≤–ª—è–µ—Ç –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –≤ Excel"""
    headers = [
        "–°—Ç–∞—Ç—É—Å –æ–±—Ä–∞–±–æ—Ç–∫–∏",
        "–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —Å—Ç–∞—Ä—Ç–∞",
        "–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —Ç–æ—á–µ–∫",
        "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ—á–µ–∫",
        "–¢–∏–ø –º–∞—Ä—à—Ä—É—Ç–∞",
        "–†–∞—Å—Å—Ç–æ—è–Ω–∏–µ 1 (–∫–º)",
        "–†–∞—Å—Å—Ç–æ—è–Ω–∏–µ 2 (–∫–º)",
        "–†–∞—Å—Å—Ç–æ—è–Ω–∏–µ 3 (–∫–º)"
    ]
    
    # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–µ–∫—É—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∫–æ–ª–æ–Ω–æ–∫
    current_max_col = ws.max_column
    
    # –ï—Å–ª–∏ —É–∂–µ –µ—Å—Ç—å –∫–æ–ª–æ–Ω–∫–∏, –Ω–∞—á–∏–Ω–∞–µ–º –ø–æ—Å–ª–µ –ø–æ—Å–ª–µ–¥–Ω–µ–π
    if current_max_col >= start_col:
        start_col = current_max_col + 1
    
    # –î–æ–±–∞–≤–ª—è–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
    for i, header in enumerate(headers):
        cell = ws.cell(row=1, column=start_col + i)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # –ù–∞—Å—Ç—Ä–∞–∏–≤–∞–µ–º —à–∏—Ä–∏–Ω—É –∫–æ–ª–æ–Ω–æ–∫
    for i in range(len(headers)):
        col_letter = get_column_letter(start_col + i)
        ws.column_dimensions[col_letter].width = 20
    
    return start_col

def validate_address_chain(address_chain):
    """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å —Ü–µ–ø–æ—á–∫–∏ –∞–¥—Ä–µ—Å–æ–≤"""
    if not address_chain:
        return False
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –¥–µ—Ñ–∏—Å–æ–≤ –¥–ª—è —Ä–∞–∑–¥–µ–ª–µ–Ω–∏—è –∞–¥—Ä–µ—Å–æ–≤
    if "-" not in address_chain:
        # –ù–æ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—Ä—è–º–æ–π –º–∞—Ä—à—Ä—É—Ç
        return True
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –∞–¥—Ä–µ—Å–∞ –Ω–µ —Å–æ–¥–µ—Ä–∂–∞—Ç –∑–∞–≤–µ–¥–æ–º–æ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
    invalid_phrases = [
        "–û—à–∏–±–∫–∞", "–æ—à–∏–±–∫–∞", "error", "Error", 
        "–ù–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–æ", "–Ω–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–æ",
        "NULL", "null", "None", "none"
    ]
    
    for phrase in invalid_phrases:
        if phrase in address_chain:
            return False
    
    return True

# ================== TELEGRAM –ë–û–¢ ==================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /start"""
    await update.message.reply_text(
        "üëã –ü—Ä–∏–≤–µ—Ç! –Ø –±–æ—Ç –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–æ–≤.\n\n"
        "üìÅ **–û—Ç–ø—Ä–∞–≤—å—Ç–µ –º–Ω–µ Excel —Ñ–∞–π–ª –≤ —Ñ–æ—Ä–º–∞—Ç–µ:**\n"
        "‚Ä¢ –ö–æ–ª–æ–Ω–∫–∞ A: –°—Ç–∞—Ä—Ç–æ–≤–∞—è —Ç–æ—á–∫–∞\n"
        "‚Ä¢ –ö–æ–ª–æ–Ω–∫–∞ B: –¶–µ–ø–æ—á–∫–∞ –∞–¥—Ä–µ—Å–æ–≤ —á–µ—Ä–µ–∑ –¥–µ—Ñ–∏—Å\n\n"
        "**–ü—Ä–∏–º–µ—Ä:**\n"
        "A1: –†–æ—Å—Ç–æ–≤-–Ω–∞-–î–æ–Ω—É, —É–ª. –û–≥–∞–Ω–æ–≤–∞ 22\n"
        "B1: –Ø—Ä–æ—Å–ª–∞–≤—Å–∫–∞—è –æ–±–ª., –≥. –†–æ—Å—Ç–æ–≤ –í–µ–ª–∏–∫–∏–π - –≥. –Ø—Ä–æ—Å–ª–∞–≤–ª—å\n\n"
        "‚úÖ –Ø –≤–µ—Ä–Ω—É —Ç–æ—Ç –∂–µ —Ñ–∞–π–ª —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ —Ä–∞—Å—á–µ—Ç–æ–≤!\n\n"
        "‚ö° –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è GraphHopper API + –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä + OpenRouteService\n"
        "üìç –ì–µ–æ–∫–æ–¥–∏—Ä—É—é—Ç—Å—è —Ç–æ–ª—å–∫–æ –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–µ –ø—É–Ω–∫—Ç—ã\n"
        "üõ£Ô∏è –†–∞—Å—á–µ—Ç –∞–≤—Ç–æ–º–æ–±–∏–ª—å–Ω—ã—Ö –º–∞—Ä—à—Ä—É—Ç–æ–≤\n\n"
        "‚ö†Ô∏è **–û–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è:**\n"
        "‚Ä¢ GraphHopper: –º–∞–∫—Å–∏–º—É–º 4 —Ç–æ—á–∫–∏ –≤ –º–∞—Ä—à—Ä—É—Ç–µ\n"
        "‚Ä¢ ORS: –¥–æ 20 —Ç–æ—á–µ–∫ (–∑–∞–ø–∞—Å–Ω–æ–π –≤–∞—Ä–∏–∞–Ω—Ç)\n"
        "‚Ä¢ –ö—Ä—ã–º, –î–ù–†, –õ–ù–† –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞—é—Ç—Å—è\n"
        "‚Ä¢ –ú–∞–ª–µ–Ω—å–∫–∏–µ –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–µ –ø—É–Ω–∫—Ç—ã –º–æ–≥—É—Ç –Ω–µ –Ω–∞–π—Ç–∏—Å—å"
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤"""
    if not update.message or not update.message.document:
        await update.message.reply_text("‚ùå –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –æ—Ç–ø—Ä–∞–≤—å—Ç–µ —Ñ–∞–π–ª")
        return
    
    file_name = update.message.document.file_name or "file.xlsx"
    file_name_lower = file_name.lower()
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ —Ñ–∞–π–ª–∞
    allowed_extensions = ['.xlsx', '.xls']
    
    if not any(file_name_lower.endswith(ext) for ext in allowed_extensions):
        await update.message.reply_text(
            "‚ùå –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –æ—Ç–ø—Ä–∞–≤—å—Ç–µ —Ñ–∞–π–ª –≤ —Ñ–æ—Ä–º–∞—Ç–µ Excel:\n"
            "‚Ä¢ .xlsx (—Ä–µ–∫–æ–º–µ–Ω–¥—É–µ—Ç—Å—è)\n"
            "‚Ä¢ .xls\n\n"
            "–ï—Å–ª–∏ —É –≤–∞—Å —Ñ–∞–π–ª –¥—Ä—É–≥–æ–≥–æ —Ñ–æ—Ä–º–∞—Ç–∞, —Å–æ—Ö—Ä–∞–Ω–∏—Ç–µ –µ–≥–æ –∫–∞–∫ .xlsx –≤ Excel."
        )
        return
    
    try:
        # –°–∫–∞—á–∏–≤–∞–µ–º —Ñ–∞–π–ª
        file = await update.message.document.get_file()
        user_id = update.message.from_user.id
        timestamp = int(time.time())
        
        # –°–æ–∑–¥–∞–µ–º –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            input_file = tmp_file.name
        
        await file.download_to_drive(input_file)
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞
        file_size = os.path.getsize(input_file)
        if file_size > 10 * 1024 * 1024:  # 10 MB
            await update.message.reply_text("‚ùå –§–∞–π–ª —Å–ª–∏—à–∫–æ–º –±–æ–ª—å—à–æ–π (–º–∞–∫—Å–∏–º—É–º 10 –ú–ë)")
            os.remove(input_file)
            return
        
        await update.message.reply_text(f"üì• –§–∞–π–ª –ø–æ–ª—É—á–µ–Ω: {file_name}")
        
        # –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
        try:
            routes, wb, ws = read_excel_with_fallback(input_file)
        except Exception as e:
            await update.message.reply_text(f"‚ùå –û—à–∏–±–∫–∞ —á—Ç–µ–Ω–∏—è —Ñ–∞–π–ª–∞: {str(e)[:200]}\n\n"
                                           "–£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ:\n"
                                           "1. –§–∞–π–ª –Ω–µ –ø–æ–≤—Ä–µ–∂–¥–µ–Ω\n"
                                           "2. –≠—Ç–æ –∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π Excel —Ñ–∞–π–ª (.xlsx)\n"
                                           "3. –î–∞–Ω–Ω—ã–µ –Ω–∞—Ö–æ–¥—è—Ç—Å—è –Ω–∞ –ø–µ—Ä–≤–æ–º –ª–∏—Å—Ç–µ\n"
                                           "4. –í –∫–æ–ª–æ–Ω–∫–µ A - —Å—Ç–∞—Ä—Ç–æ–≤—ã–µ —Ç–æ—á–∫–∏, –≤ B - —Ü–µ–ø–æ—á–∫–∏ –∞–¥—Ä–µ—Å–æ–≤")
            if os.path.exists(input_file):
                os.remove(input_file)
            return
        
        total = len(routes)
        
        if total == 0:
            await update.message.reply_text(
                "‚ùå –í —Ñ–∞–π–ª–µ –Ω–µ—Ç –¥–∞–Ω–Ω—ã—Ö –∏–ª–∏ –Ω–µ–ø—Ä–∞–≤–∏–ª—å–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç.\n\n"
                "–ü—Ä–æ–≤–µ—Ä—å—Ç–µ, —á—Ç–æ:\n"
                "1. –í –∫–æ–ª–æ–Ω–∫–µ A –µ—Å—Ç—å —Å—Ç–∞—Ä—Ç–æ–≤—ã–µ —Ç–æ—á–∫–∏\n"
                "2. –í –∫–æ–ª–æ–Ω–∫–µ B –µ—Å—Ç—å —Ü–µ–ø–æ—á–∫–∏ –∞–¥—Ä–µ—Å–æ–≤\n"
                "3. –î–∞–Ω–Ω—ã–µ –Ω–∞—á–∏–Ω–∞—é—Ç—Å—è —Å –ø–µ—Ä–≤–æ–π —Å—Ç—Ä–æ–∫–∏ (–∏–ª–∏ —Å–æ –≤—Ç–æ—Ä–æ–π, –µ—Å–ª–∏ –µ—Å—Ç—å –∑–∞–≥–æ–ª–æ–≤–∫–∏)\n"
                "4. –ê–¥—Ä–µ—Å–∞ –≤ –∫–æ–ª–æ–Ω–∫–µ B —Ä–∞–∑–¥–µ–ª–µ–Ω—ã –¥–µ—Ñ–∏—Å–æ–º (-)"
            )
            if os.path.exists(input_file):
                os.remove(input_file)
            return
        
        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–∞—á–∞–ª—å–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
        progress_msg = await update.message.reply_text(
            f"‚è≥ –ù–∞—á–∏–Ω–∞—é –æ–±—Ä–∞–±–æ—Ç–∫—É...\n"
            f"üìä –í—Å–µ–≥–æ —Å—Ç—Ä–æ–∫: {total}\n"
            f"üîë API: GraphHopper{' + –Ø–Ω–¥–µ–∫—Å' if USE_YANDEX_GEOCODER else ''}{' + ORS' if USE_ORS_FALLBACK else ''}\n"
            f"‚è±Ô∏è –û—Ä–∏–µ–Ω—Ç–∏—Ä–æ–≤–æ—á–Ω–æ–µ –≤—Ä–µ–º—è: {total * 3} —Å–µ–∫—É–Ω–¥\n\n"
            f"‚ö†Ô∏è **–í–Ω–∏–º–∞–Ω–∏–µ:**\n"
            f"‚Ä¢ GraphHopper: –º–∞–∫—Å–∏–º—É–º 4 —Ç–æ—á–∫–∏\n"
            f"‚Ä¢ ORS: –¥–æ 20 —Ç–æ—á–µ–∫ (–∑–∞–ø–∞—Å–Ω–æ–π –≤–∞—Ä–∏–∞–Ω—Ç)\n"
            f"‚Ä¢ –ö—Ä—ã–º, –î–ù–†, –õ–ù–† –ø—Ä–æ–ø—É—Å–∫–∞—é—Ç—Å—è\n"
            f"‚Ä¢ –ü–∞—É–∑—ã –º–µ–∂–¥—É –∑–∞–ø—Ä–æ—Å–∞–º–∏ –¥–ª—è API"
        )
        
        # –û—á–∏—â–∞–µ–º —Å—Ç–∞—Ä—ã–π –∫—ç—à –ø–µ—Ä–µ–¥ –Ω–∞—á–∞–ª–æ–º –æ–±—Ä–∞–±–æ—Ç–∫–∏
        print("üßπ –û—á–∏—â–∞—é —Å—Ç–∞—Ä—ã–π –∫—ç—à...")
        if os.path.exists(GEOCODE_CACHE_FILE):
            try:
                os.remove(GEOCODE_CACHE_FILE)
                print("‚úÖ –°—Ç–∞—Ä—ã–π –∫—ç—à —É–¥–∞–ª–µ–Ω")
            except:
                print("‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å —Å—Ç–∞—Ä—ã–π –∫—ç—à")
        
        # –ó–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è
        geocode_cache = load_geocode_cache()
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤
        start_col = add_result_columns(ws, start_col=3)
        
        # –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏
        processed = 0
        errors = 0
        geocode_errors = 0
        route_errors = 0
        successful = 0
        skipped = 0
        
        # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –∫–∞–∂–¥—É—é —Å—Ç—Ä–æ–∫—É
        for route in routes:
            try:
                row_num = route['row_num']
                start_point = route['start_point']
                address_chain = route['address_chain']
                
                print(f"\n{'='*60}")
                print(f"üìù –°—Ç—Ä–æ–∫–∞ {row_num}/{total}")
                print(f"üèÅ –°—Ç–∞—Ä—Ç: {start_point[:50]}...")
                print(f"üõ£Ô∏è –ú–∞—Ä—à—Ä—É—Ç: {address_chain[:50]}...")
                
                # ===== –ü–†–û–í–ï–†–ö–ê –î–ê–ù–ù–´–• =====
                if not validate_address_chain(address_chain):
                    print(f"‚ùå –ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –∞–¥—Ä–µ—Å–æ–≤, –ø—Ä–æ–ø—É—Å–∫–∞—é")
                    skipped += 1
                    
                    ws.cell(row=row_num, column=start_col).value = "‚ùå –ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –∞–¥—Ä–µ—Å–æ–≤"
                    ws.cell(row=row_num, column=start_col+1).value = "–ü—Ä–æ–ø—É—â–µ–Ω–æ"
                    ws.cell(row=row_num, column=start_col+2).value = "–ü—Ä–æ–ø—É—â–µ–Ω–æ"
                    ws.cell(row=row_num, column=start_col+3).value = 0
                    ws.cell(row=row_num, column=start_col+4).value = "–û—à–∏–±–∫–∞"
                    ws.cell(row=row_num, column=start_col+5).value = "–ü—Ä–æ–ø—É—â–µ–Ω–æ"
                    
                    processed += 1
                    continue
                
                # ===== –ì–ï–û–ö–û–î–ò–†–û–í–ê–ù–ò–ï –°–¢–ê–†–¢–û–í–û–ô –¢–û–ß–ö–ò =====
                print(f"üìç –ì–µ–æ–∫–æ–¥–∏—Ä—É—é —Å—Ç–∞—Ä—Ç–æ–≤—É—é —Ç–æ—á–∫—É...")
                start_coords = geocode_start_point(start_point)
                
                if not start_coords:
                    print(f"‚ùå –û—à–∏–±–∫–∞ –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è —Å—Ç–∞—Ä—Ç–∞: {start_point}")
                    geocode_errors += 1
                    errors += 1
                    
                    # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –æ—à–∏–±–∫—É
                    ws.cell(row=row_num, column=start_col).value = "‚ùå –û—à–∏–±–∫–∞ –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è —Å—Ç–∞—Ä—Ç–∞"
                    ws.cell(row=row_num, column=start_col+1).value = "–û—à–∏–±–∫–∞"
                    ws.cell(row=row_num, column=start_col+2).value = "–û—à–∏–±–∫–∞"
                    ws.cell(row=row_num, column=start_col+3).value = 0
                    ws.cell(row=row_num, column=start_col+4).value = "–û—à–∏–±–∫–∞"
                    ws.cell(row=row_num, column=start_col+5).value = "–û—à–∏–±–∫–∞"
                    
                    processed += 1
                    continue
                
                # ===== –ü–ê–†–°–ò–ù–ì –¶–ï–ü–û–ß–ö–ò –ê–î–†–ï–°–û–í =====
                print(f"üìç –ü–∞—Ä—Å–∏–Ω–≥ —Ü–µ–ø–æ—á–∫–∏ –∞–¥—Ä–µ—Å–æ–≤...")
                # –ò–∑–≤–ª–µ–∫–∞–µ–º —Ä–µ–≥–∏–æ–Ω –∏–∑ –ø–µ—Ä–≤–æ–≥–æ –∞–¥—Ä–µ—Å–∞ —Ü–µ–ø–æ—á–∫–∏
                first_address_region = None
                if address_chain and '-' in address_chain:
                    first_part = address_chain.split('-')[0].strip()
                    first_address_region = extract_region_from_address_improved(first_part)
                
                addresses = parse_address_chain(address_chain, first_address_region)
                
                if not addresses:
                    # –ü—Ä–æ–±—É–µ–º –∞–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω—ã–π –º–µ—Ç–æ–¥
                    addresses = extract_all_addresses_from_chain(address_chain)
                
                if not addresses:
                    print(f"‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–∞—Ä—Å–∏—Ç—å —Ü–µ–ø–æ—á–∫—É –∞–¥—Ä–µ—Å–æ–≤")
                    errors += 1
                    
                    ws.cell(row=row_num, column=start_col).value = "‚ùå –û—à–∏–±–∫–∞ –ø–∞—Ä—Å–∏–Ω–≥–∞ –∞–¥—Ä–µ—Å–æ–≤"
                    ws.cell(row=row_num, column=start_col+1).value = f"{start_coords[0]:.6f},{start_coords[1]:.6f}"
                    ws.cell(row=row_num, column=start_col+2).value = "–û—à–∏–±–∫–∞"
                    ws.cell(row=row_num, column=start_col+3).value = 0
                    ws.cell(row=row_num, column=start_col+4).value = "–û—à–∏–±–∫–∞"
                    ws.cell(row=row_num, column=start_col+5).value = "–û—à–∏–±–∫–∞"
                    
                    processed += 1
                    continue
                
                # ===== –ì–ï–û–ö–û–î–ò–†–û–í–ê–ù–ò–ï –¢–û–ß–ï–ö –ú–ê–†–®–†–£–¢–ê =====
                print(f"üìç –ì–µ–æ–∫–æ–¥–∏—Ä—É—é —Ç–æ—á–∫–∏ –º–∞—Ä—à—Ä—É—Ç–∞ ({len(addresses)} —Ç–æ—á–µ–∫)...")
                all_coords = []
                all_coords_str = []
                has_geocode_error = False
                
                for i, addr in enumerate(addresses):
                    print(f"  üìç –¢–æ—á–∫–∞ {i+1}/{len(addresses)}: {addr[:40]}...")
                    coords = enhanced_geocode(addr, geocode_cache)
                    time.sleep(0.3)  # –ü–∞—É–∑–∞ –º–µ–∂–¥—É –∑–∞–ø—Ä–æ—Å–∞–º–∏
                    
                    if coords:
                        all_coords.append(coords)
                        all_coords_str.append(f"{coords[0]:.6f},{coords[1]:.6f}")
                        print(f"    ‚úÖ –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã: {coords}")
                    else:
                        print(f"    ‚ö†Ô∏è –¢–æ—á–∫–∞ {i+1} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞, –ø—ã—Ç–∞—é—Å—å –∞–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω—ã–π –º–µ—Ç–æ–¥...")
                        
                        # –ü—Ä–æ–±—É–µ–º –∏–∑–≤–ª–µ—á—å —Ç–æ–ª—å–∫–æ –≥–æ—Ä–æ–¥
                        settlement = extract_settlement_from_address(addr)
                        if settlement:
                            simple_addr = f"{settlement}, –†–æ—Å—Å–∏—è"
                            coords = enhanced_geocode(simple_addr, geocode_cache)
                        
                        if coords:
                            all_coords.append(coords)
                            all_coords_str.append(f"{coords[0]:.6f},{coords[1]:.6f}")
                            print(f"    ‚úÖ –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —á–µ—Ä–µ–∑ —É–ø—Ä–æ—â–µ–Ω–∏–µ: {coords}")
                        else:
                            print(f"    ‚ùå –¢–æ—á–∫–∞ {i+1} –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∞, –ø—Ä–æ–ø—É—Å–∫–∞—é –º–∞—Ä—à—Ä—É—Ç")
                            has_geocode_error = True
                            geocode_errors += 1
                            break
                
                if has_geocode_error or not all_coords:
                    errors += 1
                    
                    status = "‚ùå –û—à–∏–±–∫–∞ –≥–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏—è —Ç–æ—á–µ–∫"
                    if not all_coords_str:
                        coordinates_str = "–û—à–∏–±–∫–∞"
                    else:
                        coordinates_str = "; ".join(all_coords_str)
                    
                    ws.cell(row=row_num, column=start_col).value = status
                    ws.cell(row=row_num, column=start_col+1).value = f"{start_coords[0]:.6f},{start_coords[1]:.6f}"
                    ws.cell(row=row_num, column=start_col+2).value = coordinates_str
                    ws.cell(row=row_num, column=start_col+3).value = len(addresses)
                    ws.cell(row=row_num, column=start_col+4).value = "–° –ø—Ä–æ–º–µ–∂—É—Ç–æ—á–Ω—ã–º–∏ —Ç–æ—á–∫–∞–º–∏" if len(addresses) > 1 else "–ü—Ä—è–º–æ–π"
                    ws.cell(row=row_num, column=start_col+5).value = "–û—à–∏–±–∫–∞"
                    
                    processed += 1
                    continue
                
                # ===== –†–ê–°–ß–ï–¢ –ú–ê–†–®–†–£–¢–ê =====
                route_type = "–° –ø—Ä–æ–º–µ–∂—É—Ç–æ—á–Ω—ã–º–∏ —Ç–æ—á–∫–∞–º–∏" if len(addresses) > 1 else "–ü—Ä—è–º–æ–π"
                full_coordinates = [start_coords] + all_coords
                
                # –ï—Å–ª–∏ —Ç–æ—á–µ–∫ –±–æ–ª—å—à–µ 4, –ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞–µ–º
                if len(full_coordinates) > 4:
                    print(f"‚ö†Ô∏è –í–Ω–∏–º–∞–Ω–∏–µ: {len(full_coordinates)} —Ç–æ—á–µ–∫ –≤ –º–∞—Ä—à—Ä—É—Ç–µ")
                    if len(full_coordinates) > 20:
                        route_type = f"{route_type} (—É–ø—Ä–æ—â–µ–Ω–æ –¥–æ –∫–ª—é—á–µ–≤—ã—Ö —Ç–æ—á–µ–∫)"
                    elif len(full_coordinates) > 4:
                        route_type = f"{route_type} (—Å–µ–≥–º–µ–Ω—Ç–∏—Ä–æ–≤–∞–Ω–Ω—ã–π —Ä–∞—Å—á–µ—Ç)"
                
                print(f"üìç –°—Ç—Ä–æ—é –º–∞—Ä—à—Ä—É—Ç —á–µ—Ä–µ–∑ {len(full_coordinates)} —Ç–æ—á–µ–∫...")
                
                distance = calculate_route(full_coordinates)
                time.sleep(0.5)  # –ü–∞—É–∑–∞ –¥–ª—è API
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å —Ä–∞—Å—Å—Ç–æ—è–Ω–∏—è
                if distance and distance > 0:
                    if not validate_route_distance(distance, full_coordinates):
                        print(f"‚ö†Ô∏è –ü–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ: {distance} –∫–º")
                        route_errors += 1
                        errors += 1
                        
                        ws.cell(row=row_num, column=start_col).value = "‚ö†Ô∏è –û—à–∏–±–∫–∞ —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–∞ (–ø–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ)"
                        ws.cell(row=row_num, column=start_col+1).value = f"{start_coords[0]:.6f},{start_coords[1]:.6f}"
                        ws.cell(row=row_num, column=start_col+2).value = "; ".join(all_coords_str)
                        ws.cell(row=row_num, column=start_col+3).value = len(addresses)
                        ws.cell(row=row_num, column=start_col+4).value = route_type
                        ws.cell(row=row_num, column=start_col+5).value = "–û—à–∏–±–∫–∞"
                        
                        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–∞ (–ø–æ–¥–æ–∑—Ä–∏—Ç–µ–ª—å–Ω–æ–µ —Ä–∞—Å—Å—Ç–æ—è–Ω–∏–µ)")
                    else:
                        d2, d3 = smart_variations(distance)
                        
                        # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º —É—Å–ø–µ—à–Ω—ã–π —Ä–µ–∑—É–ª—å—Ç–∞—Ç
                        ws.cell(row=row_num, column=start_col).value = "‚úÖ –£—Å–ø–µ—à–Ω–æ"
                        ws.cell(row=row_num, column=start_col+1).value = f"{start_coords[0]:.6f},{start_coords[1]:.6f}"
                        ws.cell(row=row_num, column=start_col+2).value = "; ".join(all_coords_str)
                        ws.cell(row=row_num, column=start_col+3).value = len(addresses)
                        ws.cell(row=row_num, column=start_col+4).value = route_type
                        ws.cell(row=row_num, column=start_col+5).value = distance
                        ws.cell(row=row_num, column=start_col+6).value = d2 if d2 else ""
                        ws.cell(row=row_num, column=start_col+7).value = d3 if d3 else ""
                        
                        successful += 1
                        print(f"‚úÖ –£—Å–ø–µ—à–Ω–æ: {distance} –∫–º")
                else:
                    route_errors += 1
                    errors += 1
                    
                    status = "‚ö†Ô∏è –û—à–∏–±–∫–∞ —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–∞"
                    if len(full_coordinates) > 20:
                        status = "‚ö†Ô∏è –°–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ —Ç–æ—á–µ–∫ (>20)"
                    elif len(full_coordinates) > 4:
                        status = "‚ö†Ô∏è –°–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ —Ç–æ—á–µ–∫ (>4)"
                    
                    ws.cell(row=row_num, column=start_col).value = status
                    ws.cell(row=row_num, column=start_col+1).value = f"{start_coords[0]:.6f},{start_coords[1]:.6f}"
                    ws.cell(row=row_num, column=start_col+2).value = "; ".join(all_coords_str)
                    ws.cell(row=row_num, column=start_col+3).value = len(addresses)
                    ws.cell(row=row_num, column=start_col+4).value = route_type
                    ws.cell(row=row_num, column=start_col+5).value = "–û—à–∏–±–∫–∞"
                    
                    print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–∞")
                
                processed += 1
                
                # ===== –û–ë–ù–û–í–õ–ï–ù–ò–ï –ü–†–û–ì–†–ï–°–°–ê =====
                if processed % 2 == 0 or processed == total:
                    try:
                        progress_percent = int((processed / total) * 100)
                        
                        progress_text = (
                            f"‚è≥ –û–±—Ä–∞–±–æ—Ç–∫–∞: {processed}/{total} ({progress_percent}%)\n"
                            f"‚úÖ –£—Å–ø–µ—à–Ω–æ: {successful}\n"
                            f"‚ùå –û—à–∏–±–∫–∏: {errors}\n"
                            f"‚è≠Ô∏è –ü—Ä–æ–ø—É—â–µ–Ω–æ: {skipped}\n"
                        )
                        
                        if geocode_errors > 0:
                            progress_text += f"üìç –ì–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ: {geocode_errors}\n"
                        
                        if route_errors > 0:
                            progress_text += f"üõ£Ô∏è –ú–∞—Ä—à—Ä—É—Ç—ã: {route_errors}\n"
                        
                        # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º —Ç–µ–∫—É—â–∏–π –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º—ã–π –≥–æ—Ä–æ–¥
                        if processed < total and successful > 0:
                            settlement = extract_settlement_from_address(start_point)
                            if settlement:
                                progress_text += f"üìç –¢–µ–∫—É—â–∏–π: {settlement[:30]}..."
                        
                        await progress_msg.edit_text(progress_text)
                    except Exception as e:
                        print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞: {e}")
                        
            except Exception as e:
                print(f"‚ùå –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞ –≤ —Å—Ç—Ä–æ–∫–µ {row_num}: {e}")
                log_error(row_num, f"{start_point[:50]}...", "CRITICAL", str(e))
                errors += 1
                processed += 1
        
        # ===== –°–û–•–†–ê–ù–ï–ù–ò–ï –ö–≠–®–ê =====
        save_geocode_cache(geocode_cache)
        
        # ===== –°–û–•–†–ê–ù–ï–ù–ò–ï –ò –û–¢–ü–†–ê–í–ö–ê –†–ï–ó–£–õ–¨–¢–ê–¢–ê =====
        try:
            await progress_msg.edit_text(
                f"‚úÖ –û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞!\n"
                f"üìä –ò—Ç–æ–≥–∏:\n"
                f"‚Ä¢ –í—Å–µ–≥–æ —Å—Ç—Ä–æ–∫: {total}\n"
                f"‚Ä¢ –£—Å–ø–µ—à–Ω–æ: {successful}\n"
                f"‚Ä¢ –û—à–∏–±–æ–∫: {errors}\n"
                f"‚Ä¢ –ü—Ä–æ–ø—É—â–µ–Ω–æ: {skipped}\n"
                f"  ‚îî –ì–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ: {geocode_errors}\n"
                f"  ‚îî –†–∞—Å—á–µ—Ç –º–∞—Ä—à—Ä—É—Ç–æ–≤: {route_errors}\n\n"
                f"üíæ –°–æ—Ö—Ä–∞–Ω—è—é —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã..."
            )
        except:
            pass
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç
        output_file = f"results_{user_id}_{timestamp}.xlsx"
        wb.save(output_file)
        
        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç
        try:
            with open(output_file, "rb") as file:
                caption = (
                    f"‚úÖ –û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞!\n\n"
                    f"üìä **–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞:**\n"
                    f"‚Ä¢ –í—Å–µ–≥–æ —Å—Ç—Ä–æ–∫: {total}\n"
                    f"‚Ä¢ –£—Å–ø–µ—à–Ω–æ: {successful}\n"
                    f"‚Ä¢ –û—à–∏–±–æ–∫: {errors}\n"
                    f"‚Ä¢ –ü—Ä–æ–ø—É—â–µ–Ω–æ: {skipped}\n\n"
                    f"‚ö° **–ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–æ:**\n"
                    f"‚Ä¢ GraphHopper API\n"
                    f"‚Ä¢ –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä\n"
                    f"‚Ä¢ ORS API (–∑–∞–ø–∞—Å–Ω–æ–π)\n"
                    f"‚Ä¢ –ì–µ–æ–∫–æ–¥–∏—Ä–æ–≤–∞–Ω–∏–µ –ø–æ –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–º –ø—É–Ω–∫—Ç–∞–º\n"
                    f"‚Ä¢ –†–∞—Å—á–µ—Ç –∞–≤—Ç–æ–º–æ–±–∏–ª—å–Ω—ã—Ö –º–∞—Ä—à—Ä—É—Ç–æ–≤\n\n"
                    f"‚ö†Ô∏è **–û–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è:**\n"
                    f"‚Ä¢ GraphHopper: –º–∞–∫—Å–∏–º—É–º 4 —Ç–æ—á–∫–∏\n"
                    f"‚Ä¢ ORS: –¥–æ 20 —Ç–æ—á–µ–∫ (–∑–∞–ø–∞—Å–Ω–æ–π)\n"
                    f"‚Ä¢ –ö—Ä—ã–º, –î–ù–†, –õ–ù–† –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞—é—Ç—Å—è\n\n"
                    f"üìé –§–∞–π–ª: {file_name}"
                )
                
                await update.message.reply_document(
                    document=file,
                    filename=f"—Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã_{file_name}",
                    caption=caption,
                    parse_mode='Markdown'
                )
            
            print(f"‚úÖ –§–∞–π–ª –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {user_id}")
            
        except Exception as e:
            await update.message.reply_text(f"‚ùå –û—à–∏–±–∫–∞ –æ—Ç–ø—Ä–∞–≤–∫–∏ —Ñ–∞–π–ª–∞: {str(e)[:200]}")
        
        # ===== –û–ß–ò–°–¢–ö–ê =====
        try:
            if os.path.exists(input_file):
                os.remove(input_file)
            if os.path.exists(output_file):
                os.remove(output_file)
        except Exception as e:
            print(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –æ—á–∏—Å—Ç–∫–∏ —Ñ–∞–π–ª–æ–≤: {e}")
        
    except Exception as e:
        error_msg = str(e)[:500]
        await update.message.reply_text(f"‚ùå –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞: {error_msg}\n\n"
                                       "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–ø—Ä–æ–±—É–π—Ç–µ:\n"
                                       "1. –°–æ—Ö—Ä–∞–Ω–∏—Ç—å —Ñ–∞–π–ª –∫–∞–∫ .xlsx –≤ Excel\n"
                                       "2. –ü—Ä–æ–≤–µ—Ä–∏—Ç—å, —á—Ç–æ —Ñ–∞–π–ª –Ω–µ –ø–æ–≤—Ä–µ–∂–¥–µ–Ω\n"
                                       "3. –û—Ç–ø—Ä–∞–≤–∏—Ç—å —Ñ–∞–π–ª –∑–∞–Ω–æ–≤–æ")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /help"""
    help_text = """
üìã **–î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–º–∞–Ω–¥—ã:**

/start - –ù–∞—á–∞–ª–æ —Ä–∞–±–æ—Ç—ã
/help - –≠—Ç–∞ —Å–ø—Ä–∞–≤–∫–∞

üìÅ **–§–æ—Ä–º–∞—Ç Excel —Ñ–∞–π–ª–∞:**

| –ö–æ–ª–æ–Ω–∫–∞ A | –ö–æ–ª–æ–Ω–∫–∞ B |
|-----------|-----------|
| –°—Ç–∞—Ä—Ç–æ–≤–∞—è —Ç–æ—á–∫–∞ | –¶–µ–ø–æ—á–∫–∞ –∞–¥—Ä–µ—Å–æ–≤ —á–µ—Ä–µ–∑ –¥–µ—Ñ–∏—Å |

üìç **–ü—Ä–∏–º–µ—Ä –¥–∞–Ω–Ω—ã—Ö:**
A1: –†–æ—Å—Ç–æ–≤-–Ω–∞-–î–æ–Ω—É, –û–≥–∞–Ω–æ–≤–∞ 22
B1: –Ø—Ä–æ—Å–ª–∞–≤—Å–∫–∞—è –æ–±–ª., –≥. –†–æ—Å—Ç–æ–≤ –í–µ–ª–∏–∫–∏–π - –≥. –Ø—Ä–æ—Å–ª–∞–≤–ª—å

üìä **–î–æ–±–∞–≤–ª—è–µ–º—ã–µ –∫–æ–ª–æ–Ω–∫–∏:**
1. –°—Ç–∞—Ç—É—Å –æ–±—Ä–∞–±–æ—Ç–∫–∏
2. –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —Å—Ç–∞—Ä—Ç–∞
3. –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã —Ç–æ—á–µ–∫
4. –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ—á–µ–∫
5. –¢–∏–ø –º–∞—Ä—à—Ä—É—Ç–∞
6. –†–∞—Å—Å—Ç–æ—è–Ω–∏–µ 1 (–∫–º)
7. –†–∞—Å—Å—Ç–æ—è–Ω–∏–µ 2 (–∫–º)
8. –†–∞—Å—Å—Ç–æ—è–Ω–∏–µ 3 (–∫–º)

‚ö° **–û—Å–æ–±–µ–Ω–Ω–æ—Å—Ç–∏:**
‚Ä¢ –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è GraphHopper API + –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä + OpenRouteService
‚Ä¢ –ì–µ–æ–∫–æ–¥–∏—Ä—É—é—Ç—Å—è —Ç–æ–ª—å–∫–æ –≥–æ—Ä–æ–¥–∞/–Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–µ –ø—É–Ω–∫—Ç—ã
‚Ä¢ –£–ª–∏—Ü—ã –∏ –Ω–æ–º–µ—Ä–∞ –¥–æ–º–æ–≤ –∏–≥–Ω–æ—Ä–∏—Ä—É—é—Ç—Å—è
‚Ä¢ –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–æ–µ –ø—Ä–∏–º–µ–Ω–µ–Ω–∏–µ —Ä–µ–≥–∏–æ–Ω–æ–≤

‚ö†Ô∏è **–û–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è:**
‚Ä¢ GraphHopper: –º–∞–∫—Å–∏–º—É–º 4 —Ç–æ—á–∫–∏ –≤ –º–∞—Ä—à—Ä—É—Ç–µ
‚Ä¢ ORS: –¥–æ 20 —Ç–æ—á–µ–∫ (–∑–∞–ø–∞—Å–Ω–æ–π –≤–∞—Ä–∏–∞–Ω—Ç)
‚Ä¢ –ö—Ä—ã–º, –î–ù–†, –õ–ù–†, –•–µ—Ä—Å–æ–Ω—Å–∫–∞—è, –ó–∞–ø–æ—Ä–æ–∂—Å–∫–∞—è –æ–±–ª–∞—Å—Ç–∏ –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞—é—Ç—Å—è
‚Ä¢ –ú–∞–ª–µ–Ω—å–∫–∏–µ –Ω–∞—Å–µ–ª–µ–Ω–Ω—ã–µ –ø—É–Ω–∫—Ç—ã –º–æ–≥—É—Ç –Ω–µ –Ω–∞–π—Ç–∏—Å—å
‚Ä¢ –ü–∞—É–∑—ã –º–µ–∂–¥—É –∑–∞–ø—Ä–æ—Å–∞–º–∏ –¥–ª—è —Å–æ–±–ª—é–¥–µ–Ω–∏—è –ª–∏–º–∏—Ç–æ–≤ API
"""
    await update.message.reply_text(help_text)

async def test_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """–¢–µ—Å—Ç–æ–≤–∞—è –∫–æ–º–∞–Ω–¥–∞ –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ä–∞–±–æ—Ç—ã –±–æ—Ç–∞"""
    api_status = "‚úÖ –î–æ—Å—Ç—É–ø–µ–Ω" if GRAPHHOPPER_API_KEY else "‚ùå –ù–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω"
    yandex_status = "‚úÖ –ù–∞—Å—Ç—Ä–æ–µ–Ω" if YANDEX_GEOCODER_API_KEY else "‚ùå –ù–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω"
    ors_status = "‚úÖ –ù–∞—Å—Ç—Ä–æ–µ–Ω" if ORS_API_KEY else "‚ùå –ù–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω"
    
    await update.message.reply_text(
        f"ü§ñ –ë–æ—Ç —Ä–∞–±–æ—Ç–∞–µ—Ç!\n\n"
        f"–û—Ç–ø—Ä–∞–≤—å—Ç–µ Excel —Ñ–∞–π–ª –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –º–∞—Ä—à—Ä—É—Ç–æ–≤.\n\n"
        f"GraphHopper API: {api_status}\n"
        f"–Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä: {yandex_status}\n"
        f"OpenRouteService: {ors_status}\n\n"
        f"‚ö†Ô∏è –î–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä API –∫–ª—é—á–∞:\n"
        f"1. –ó–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–π—Ç–µ—Å—å –Ω–∞ https://developer.tech.yandex.ru/\n"
        f"2. –ü–æ–ª—É—á–∏—Ç–µ API –∫–ª—é—á –¥–ª—è –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä–∞\n"
        f"3. –î–æ–±–∞–≤—å—Ç–µ –ø–µ—Ä–µ–º–µ–Ω–Ω—É—é YANDEX_GEOCODER_API_KEY –≤ Render"
    )

# ================== –ó–ê–ü–£–°–ö –ë–û–¢–ê ==================
async def run_bot():
    """–ó–∞–ø—É—Å–∫–∞–µ—Ç –±–æ—Ç–∞ —Å –æ–±—Ä–∞–±–æ—Ç–∫–æ–π –∫–æ–Ω—Ñ–ª–∏–∫—Ç–æ–≤"""
    print("=" * 60)
    print("üöÄ –ó–ê–ü–£–°–ö –¢–ï–õ–ï–ì–†–ê–ú –ë–û–¢–ê")
    print("=" * 60)
    
    if not BOT_TOKEN:
        print("‚ùå –û–®–ò–ë–ö–ê: BOT_TOKEN –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω!")
        print("–£—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ –ø–µ—Ä–µ–º–µ–Ω–Ω—É—é –æ–∫—Ä—É–∂–µ–Ω–∏—è BOT_TOKEN –≤ Render")
        return
    
    print(f"‚úÖ –¢–æ–∫–µ–Ω –ø–æ–ª—É—á–µ–Ω")
    print(f"‚úÖ GraphHopper API –∫–ª—é—á: {'‚úÖ –ù–∞—Å—Ç—Ä–æ–µ–Ω' if GRAPHHOPPER_API_KEY else '‚ùå –ù–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω'}")
    print(f"‚úÖ –Ø–Ω–¥–µ–∫—Å.–ì–µ–æ–∫–æ–¥–µ—Ä API –∫–ª—é—á: {'‚úÖ –ù–∞—Å—Ç—Ä–æ–µ–Ω' if YANDEX_GEOCODER_API_KEY else '‚ùå –ù–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω'}")
    print(f"‚úÖ OpenRouteService API –∫–ª—é—á: {'‚úÖ –ù–∞—Å—Ç—Ä–æ–µ–Ω' if ORS_API_KEY else '‚ùå –ù–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω'}")
    
    if not GRAPHHOPPER_API_KEY:
        print("‚ö†Ô∏è –í–ù–ò–ú–ê–ù–ò–ï: GraphHopper API –∫–ª—é—á –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω!")
        print("–î–æ–±–∞–≤—å—Ç–µ –ø–µ—Ä–µ–º–µ–Ω–Ω—É—é GRAPHHOPPER_API_KEY –≤ Render")
    
    # –°–æ–∑–¥–∞–µ–º –ø—Ä–∏–ª–æ–∂–µ–Ω–∏–µ
    application = ApplicationBuilder().token(BOT_TOKEN).build()
    
    # –î–æ–±–∞–≤–ª—è–µ–º –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("test", test_command))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    # –ü—ã—Ç–∞–µ–º—Å—è –∑–∞–ø—É—Å—Ç–∏—Ç—å –±–æ—Ç–∞
    max_retries = 5
    retry_delay = 10
    
    for attempt in range(max_retries):
        try:
            print(f"üîÑ –ü–æ–ø—ã—Ç–∫–∞ {attempt + 1}/{max_retries} –∑–∞–ø—É—Å—Ç–∏—Ç—å –±–æ—Ç–∞...")
            await application.initialize()
            await application.start()
            
            # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –±–æ—Ç–µ
            bot_info = await application.bot.get_me()
            print(f"‚úÖ –ë–æ—Ç –∑–∞–ø—É—â–µ–Ω: @{bot_info.username}")
            
            # –ó–∞–ø—É—Å–∫–∞–µ–º polling
            await application.updater.start_polling(
                drop_pending_updates=True,
                timeout=30,
                poll_interval=0.5
            )
            
            print("ü§ñ –ë–æ—Ç —Ä–∞–±–æ—Ç–∞–µ—Ç –∏ –æ–∂–∏–¥–∞–µ—Ç —Å–æ–æ–±—â–µ–Ω–∏–π...")
            print("‚ÑπÔ∏è –î–ª—è –æ—Å—Ç–∞–Ω–æ–≤–∫–∏ –Ω–∞–∂–º–∏—Ç–µ Ctrl+C")
            
            # –ë–µ—Å–∫–æ–Ω–µ—á–Ω—ã–π —Ü–∏–∫–ª
            while True:
                await asyncio.sleep(3600)
            
        except Conflict as e:
            print(f"‚ö†Ô∏è –ö–æ–Ω—Ñ–ª–∏–∫—Ç: {e}")
            print(f"‚è≥ –ñ–¥—É {retry_delay} —Å–µ–∫—É–Ω–¥ –ø–µ—Ä–µ–¥ –ø–æ–≤—Ç–æ—Ä–Ω–æ–π –ø–æ–ø—ã—Ç–∫–æ–π...")
            
            try:
                await application.stop()
                await application.shutdown()
            except:
                pass
            
            if attempt < max_retries - 1:
                await asyncio.sleep(retry_delay)
                retry_delay *= 2
            else:
                print("‚ùå –î–æ—Å—Ç–∏–≥–Ω—É—Ç –ª–∏–º–∏—Ç –ø–æ–ø—ã—Ç–æ–∫. –ë–æ—Ç –Ω–µ –º–æ–∂–µ—Ç –∑–∞–ø—É—Å—Ç–∏—Ç—å—Å—è.")
                break
                
        except Exception as e:
            print(f"‚ùå –û—à–∏–±–∫–∞: {e}")
            break

def main():
    """–û—Å–Ω–æ–≤–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –∑–∞–ø—É—Å–∫–∞"""
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Ä–∞–±–æ—Ç–∞–µ–º –ª–∏ –Ω–∞ Render
    is_render = os.environ.get('RENDER') is not None
    port = os.environ.get('PORT')
    
    if is_render and port:
        print(f"üåê –†–∞–±–æ—Ç–∞–µ–º –Ω–∞ Render, –ø–æ—Ä—Ç: {port}")
        # –ó–∞–ø—É—Å–∫–∞–µ–º Flask –≤ –æ—Ç–¥–µ–ª—å–Ω–æ–º –ø–æ—Ç–æ–∫–µ
        flask_thread = threading.Thread(target=run_flask, daemon=True)
        flask_thread.start()
        print("‚úÖ Flask —Å–µ—Ä–≤–µ—Ä –∑–∞–ø—É—â–µ–Ω –≤ –æ—Ç–¥–µ–ª—å–Ω–æ–º –ø–æ—Ç–æ–∫–µ")
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º –±–æ—Ç–∞
    try:
        asyncio.run(run_bot())
    except KeyboardInterrupt:
        print("\nüëã –ë–æ—Ç –æ—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º")
    except Exception as e:
        print(f"‚ùå –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞: {e}")

if __name__ == "__main__":
    main()